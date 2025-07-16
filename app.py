# app.py  – Hotel Labor Tool
# ───────────────────────────────────────────────────────────────
import os
import streamlit as st
import pandas as pd
import numpy as np
import datetime
#import pdfkit
from jinja2 import Template
from datetime import date, timedelta
from st_aggrid import JsCode
from sqlalchemy import func
from dateutil.relativedelta import relativedelta, MO
from sqlalchemy.orm import scoped_session
from sqlalchemy import or_, func
from sqlalchemy import create_engine, text
ENGINE = create_engine("sqlite:///hotel_labor.db", echo=False)

# ───── ONE-TIME RESET FOR OT RISK ─────
with ENGINE.connect() as conn:
    conn.execute(text("DROP TABLE IF EXISTS ot_risk_all"))
    print("✅ Dropped old ot_risk_all table with missing columns")
from st_aggrid import AgGrid, GridOptionsBuilder
from st_aggrid import AgGrid, GridOptionsBuilder
from db import Session, LaborStandard, RoomForecast
import pandas as pd
import math
import os, json, math, pandas as pd
from datetime import date, timedelta, datetime, time
from db import Position, LaborStandard
from db import RoomForecast
from datetime import date, timedelta
from collections import defaultdict
from db import Actual, Schedule, RoomActual, RoomForecast, Position, Department, ShiftTime
from db import Schedule, Employee, Position, ShiftTime, RoomForecast, Actual
import db                                # local ORM layer

import streamlit as st
import requests

API_URL = "http://localhost:8000"  # change to your FastAPI URL if deployed

# 🔐 Step 1: Login Form
if "token" not in st.session_state:
    st.title("Hotel Login")
    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submit = st.form_submit_button("Login")

    if submit:
        response = requests.post(f"{API_URL}/login", json={
            "username": username,
            "password": password
        })

        if response.status_code == 200:
            st.success("Login successful!")
            st.session_state.token = response.json()["access_token"]

            # 🔄 Fetch hotel info for filtering
            user = requests.get(f"{API_URL}/me", headers={
                "Authorization": f"Bearer {st.session_state.token}"
            })

            if user.status_code == 200:
                st.session_state.hotel_name = user.json()["hotel_name"]
                st.rerun()
            else:
                st.error("Could not retrieve user info.")
        else:
            st.error("Invalid username or password.")
    st.stop()  # stop execution until logged in

# ─────────────── Utility Functions ───────────────

def get_week_start(any_date=None):
      if any_date is None:
            any_date = date.today()
      return any_date - timedelta(days=any_date.weekday())  # Week starts on Monday

def today():
      return date.today()

def generate_ot_risk_data(week_start, week_end, sel_dept, sel_pos):
    from sqlalchemy import or_

    # ───── Pull actual hours with department & position ─────
    q = (
        session.query(
            db.Actual.emp_id.label("Number"),
            db.Actual.date.label("Business Date"),
            (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
            db.Position.name.label("Position"),
            db.Department.name.label("Department")
        )
        .join(db.Position, db.Actual.position_id == db.Position.id)
        .join(db.Department, db.Position.department_id == db.Department.id)
        .filter(db.Actual.date.between(week_start, week_end))
        .filter(or_(db.Actual.hours != 0, db.Actual.ot_hours != 0))
    )

    if sel_dept:
        q = q.filter(db.Department.name == sel_dept)
    if sel_pos:
        q = q.filter(db.Position.name == sel_pos)

    raw = pd.DataFrame(q.all(), columns=["Number", "Business Date", "Hours", "Position", "Department"])

    if raw.empty:
        return pd.DataFrame()

    # ───── Match employee names ─────
    emp_df = refresh(db.Employee).copy()
    parts = emp_df["name"].astype(str).str.extract(
        r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
    )
    emp_df["ID"] = parts["ID"].fillna("").astype(str).str.zfill(5)
    emp_df["First Name"] = parts["First_Name"].str.strip()
    emp_df["Last Name"] = parts["Last_Name"].str.strip()
    emp_df["match_ID"] = emp_df["ID"].astype(str).str.lstrip("0")
    raw["match_ID"] = raw["Number"].astype(str).str.lstrip("0")

    merged = raw.merge(emp_df[["match_ID", "First Name", "Last Name"]], on="match_ID", how="left")

    # ───── Aggregate actuals ─────
    agg = merged.groupby(["Number", "First Name", "Last Name"]).agg(
        total_hours=("Hours", "sum"),
        days_worked=("Business Date", pd.Series.nunique)
    ).reset_index()

    # Ensure 'Number' is string before merging
    agg["Number"] = agg["Number"].astype(str)

    # ───── Fill placeholders for shift logic ─────
    agg["Days Scheduled"] = 0
    agg["Days Remaining"] = 0
    agg["Future Scheduled Hrs"] = 0
    agg["Total Hrs Worked + Schedule"] = agg["total_hours"].round(2)

    def classify_ot_risk(row):
        if row["Total Hrs Worked + Schedule"] <= 40:
            return "No Risk"
        return "OT"

    agg["OT Risk"] = agg.apply(classify_ot_risk, axis=1)
    agg["OT Risk %"] = agg["total_hours"].apply(
        lambda h: "0%" if h <= 40 else f"{round(((h - 40)/40)*100)}%"
    )
    agg["Projected OT"] = agg["total_hours"].apply(lambda h: max(round(h - 40, 2), 0))

    # ───── Merge hourly rate ─────
    emp_df["ID"] = emp_df["ID"].astype(str)  # Coerce to match
    if "hourly_rate" in emp_df.columns:
        emp_df["rate"] = emp_df["hourly_rate"].fillna(0)
    else:
        emp_df["rate"] = 0.00

    agg = agg.merge(emp_df[["ID", "rate"]], left_on="Number", right_on="ID", how="left")
    agg["rate"] = agg["rate"].fillna(0)
    agg["OT Cost"] = (agg["Projected OT"] * agg["rate"] * 1.5).round(2)

    # ───── Add week, department, position for filtering ─────
    agg["department"] = sel_dept
    agg["position"] = sel_pos

    return agg
def save_ot_risk_to_db(week_start, week_end, sel_dept, sel_pos):
    from sqlalchemy import create_engine
    ENGINE = create_engine("sqlite:///hotel_labor.db", echo=False)

    df = generate_ot_risk_data(week_start, week_end, sel_dept, sel_pos)
    if df.empty:
        return False

    df["Business Date"] = pd.to_datetime(week_start)  # Required for filtering later
    with ENGINE.connect() as conn:
        try:
            df.to_sql("ot_risk_all", conn, if_exists="append", index=False)
            return True
        except Exception as e:
            print("❌ Failed to save OT Risk:", e)
            return False

session = scoped_session(db.Session)

# ---------- helpers -----------------------------------------------------------
def refresh(model):
    """Return full SQLAlchemy table as DataFrame."""
    return pd.read_sql(session.query(model).statement, session.bind)

# ---------- helper: convert weekday label -> placeholder DATE --------------
from datetime import date  # already imported at top
# map 3-letter weekday -> placeholder DATE in the dummy week
WK_TO_DATE = {
    "Mon": date(2000, 1, 3),
    "Tue": date(2000, 1, 4),
    "Wed": date(2000, 1, 5),
    "Thu": date(2000, 1, 6),
    "Fri": date(2000, 1, 7),
    "Sat": date(2000, 1, 8),
    "Sun": date(2000, 1, 9),
}
def week_cols(week_start):
    """Return 7 day headers; %#d/#m for Windows, %-d/-m for POSIX."""
    fmt = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
    return [(week_start + timedelta(d)).strftime(fmt) for d in range(7)]

def refresh_totals(session, keys):
    """
    keys → set of (position_id, date) pairs.
    Updates the 'total' source in db.Actual to equal manual + contract for each.
    """
    for pos_id, biz_date in keys:
        # Sum all relevant sources (manual + contract)
        sums = session.query(
            func.coalesce(func.sum(db.Actual.hours), 0),
            func.coalesce(func.sum(db.Actual.ot_hours), 0),
            func.coalesce(func.sum(db.Actual.reg_pay), 0),
            func.coalesce(func.sum(db.Actual.ot_pay), 0),
        ).filter(
            db.Actual.position_id == pos_id,
            db.Actual.date == biz_date,
            db.Actual.source.in_(["manual", "contract"])
        ).one()

        total_rec = (
            session.query(db.Actual)
                .filter_by(position_id=pos_id, date=biz_date, source="total")
                .one_or_none()
        )

        if not total_rec:
            total_rec = db.Actual(
                emp_id=None,
                position_id=pos_id,
                date=biz_date,
                source="total"
            )
            session.add(total_rec)

        total_rec.hours, total_rec.ot_hours, total_rec.reg_pay, total_rec.ot_pay = sums

def get_week_dates():
    from datetime import date, timedelta
    today = date.today()
    start = today - timedelta(days=today.weekday())
    return [start + timedelta(days=i) for i in range(7)]


def load_labor_standards(pos_id):
    rows = session.query(db.LaborStandard).filter_by(position_id=pos_id).all()
    return pd.DataFrame([{
        "Metric": r.metric,
        "Standard": r.standard,
        "Unit": r.unit
    } for r in rows]) if rows else pd.DataFrame(columns=["Metric", "Standard", "Unit"])

def save_labor_standards(pos_id, df):
    session.query(db.LaborStandard).filter_by(position_id=pos_id).delete()
    new_rows = [
        db.LaborStandard(position_id=pos_id, metric=row["Metric"], standard=row["Standard"], unit=row["Unit"])
        for _, row in df.iterrows() if row["Metric"] and row["Standard"]
    ]
    session.bulk_save_objects(new_rows)
    session.commit()


# ---------- Streamlit page config ---------------------------------------------
st.set_page_config("Hotel Labor Tool", layout="wide")
# ───── Cleaner Compact Title Bar ─────
st.markdown("""
    <style>
    .top-bar-title {
        font-size: 20px;
        font-weight: 600;
        padding-top: 0;
        margin-top: -25px;
        margin-bottom: 10px;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown('<div class="top-bar-title">🏨 Hotel Labor Management</div>', unsafe_allow_html=True)

main_choice = st.sidebar.radio(
    "Navigation",
    ["Dashboard",
     "Employees",
     "Labor ▸ Structure",
     "Labor ▸ Actual Hours",
     "Room STATs",
     "Scheduling",
     "Cost and OT Mgmt",
      "Reports"],
    format_func=lambda x: x.replace("Labor ▸ ", "")
)

st.markdown("""
    <style>
    /* ───── Dynamic Sidebar Gradient Background ───── */
    [data-testid="stSidebar"] {
        background: linear-gradient(135deg, #2e2e2e, #4a4a4a, #3c3c3c) !important;
        background-size: 400% 400%;
        animation: gradientShift 20s ease infinite;
        box-shadow: 2px 0 10px rgba(0,0,0,0.25);
    }

    @keyframes gradientShift {
        0%   { background-position: 0% 50%; }
        50%  { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }

    /* ───── Sidebar Width ───── */
    [data-testid="stSidebar"][aria-expanded="true"] > div:first-child {
        width: 220px;
    }

    /* ───── Sidebar Text Styling ───── */
    [data-testid="stSidebar"] * {
        color: #ffffff !important;
    }

    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] span {
        font-size: 15px !important;
    }

    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3 {
        color: #f0f0f0 !important;
    }
    </style>
""", unsafe_allow_html=True)
# ======================================= DASHBOARD ==========================
if main_choice == "Dashboard":
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    from datetime import date, timedelta
    import math, pandas as pd
    from sqlalchemy import func, and_

    st.title("📊 Labor Analytics Dashboard")

    # ─────────── 1. FILTER BAR ────────────────────────────────────────────
    f_col1, f_col2, f_col3, f_col4 = st.columns(4)

    with f_col1:
        dt_from = st.date_input(
            "From",
            value=st.session_state.get("sch_week_start", date.today()),
            key="la_from",
        )
    with f_col2:
        dt_to = st.date_input(
            "To",
            value=max(dt_from, date.today()),
            min_value=dt_from,
            key="la_to",
        )

    with f_col3:
        sel_dept = st.selectbox(
            "Department",
            ["All"]
            + sorted(refresh(db.Employee)["department"].dropna().unique()),
            key="la_dept",
        )
    with f_col4:
        pos_opts = (
            refresh(db.Employee)
            .query("department == @sel_dept")["role"]
            .dropna()
            .unique()
            if sel_dept != "All"
            else refresh(db.Employee)["role"].dropna().unique()
        )
        sel_pos = st.selectbox(
            "Position", ["All"] + sorted(pos_opts), key="la_pos"
        )

    # ─────────── 2. LIVE DATA HELPERS ────────────────────────────────────

    def get_actual_hours(dep, pos, start, end):
        q = (
            session.query(
                Actual.date.label("date"),
                func.sum(Actual.hours).label("hours"),
                func.sum(Actual.ot_hours).label("ot_hours"),
            )
            .join(Position, Position.id == Actual.position_id)
        )

        if dep != "All":
            q = q.join(
                Department, Department.id == Position.department_id
            ).filter(Department.name == dep)
        if pos != "All":
            q = q.filter(Position.name == pos)

        q = q.filter(Actual.date.between(start, end)).group_by(
            Actual.date
        )

        df = pd.read_sql(q.statement, session.bind)
        df["date"] = pd.to_datetime(df["date"])  # 🔧 Ensure datetime64[ns]

        all_days = pd.DataFrame({"date": pd.date_range(start, end)})
        all_days["date"] = pd.to_datetime(all_days["date"])  # 🔧 Ensure datetime64[ns]

        return all_days.merge(df, on="date", how="left").fillna(0)

    def get_standard_hours(dep, pos, start, end):
        q = (
            session.query(
                Schedule.day.label("date"),
                func.count().label("std_shift_count")  # count of scheduled shifts
            )
            .join(Employee, Employee.id == Schedule.emp_id)
            .join(Position, Position.name == Employee.role)
        )

        if dep != "All":
            q = (
                q.join(Department, Department.id == Position.department_id)
                 .filter(Department.name == dep)
            )
        if pos != "All":
            q = q.filter(Position.name == pos)

        q = q.filter(Schedule.day.between(start, end)).group_by(Schedule.day)

        df = pd.read_sql(q.statement, session.bind)
        df["date"] = pd.to_datetime(df["date"])
        df["standard_hours"] = df["std_shift_count"] * 8  # ← standard shift length
        df.drop(columns="std_shift_count", inplace=True)

        all_days = pd.DataFrame({"date": pd.date_range(start, end)})
        all_days["date"] = pd.to_datetime(all_days["date"])

        return all_days.merge(df, on="date", how="left").fillna(0)
    def get_schedule_hours(dep, pos, start, end):
        """
        Earned FTEs = one 8-hour FTE for every shift that appears
        on Schedule between `start` and `end`, filtered by Department / Position.
        """

        # ── base query: count shifts per calendar day ─────────────────────
        q = (
            session.query(
                Schedule.day.label("date"),
                func.count().label("shift_cnt")          # number of shifts that day
            )
            .join(Employee, Employee.id == Schedule.emp_id)
            .join(Position, Position.name == Employee.role)
        )

        if dep != "All":
            q = (
                q.join(Department, Department.id == Position.department_id)
                 .filter(Department.name == dep)
            )
        if pos != "All":
            q = q.filter(Position.name == pos)

        q = (
            q.filter(Schedule.day.between(start, end))
             .group_by(Schedule.day)
        )

        # rows → DataFrame
        df = pd.read_sql(q.statement, session.bind)
        df["date"] = pd.to_datetime(df["date"])

        # convert shift count → hours (8 h per shift) and FTEs
        df["sched_hours"] = df["shift_cnt"] * 8        # total hours scheduled
        df.drop(columns="shift_cnt", inplace=True)

        # ── ensure every day in the range exists ───────────────────────────
        all_days = pd.DataFrame({"date": pd.date_range(start, end)})
        all_days["date"] = pd.to_datetime(all_days["date"])

        return (
            all_days.merge(df, on="date", how="left")
                    .fillna(0)                         # days with no schedule → 0
        )
    def get_room_stats(metric, start, end):
        sub_act = (
            session.query(
                RoomActual.date,
                RoomActual.value.label("actual"),
            )
            .filter(RoomActual.kpi.ilike(metric))
            .filter(RoomActual.date.between(start, end))
            .subquery()
        )

        sub_fc = (
            session.query(
                RoomForecast.date,
                RoomForecast.value.label("forecast"),
            )
            .filter(RoomForecast.kpi.ilike(metric))
            .filter(RoomForecast.date.between(start, end))
            .subquery()
        )

        q = (
            session.query(
                sub_act.c.date.label("date"),
                sub_act.c.actual,
                sub_fc.c.forecast,
            )
            .select_from(
                sub_act.outerjoin(sub_fc, sub_act.c.date == sub_fc.c.date)
            )
            .order_by(sub_act.c.date)
        )

        df = pd.read_sql(q.statement, session.bind)
        df["date"] = pd.to_datetime(df["date"])  # 🔧 Ensure datetime64[ns]

        all_days = pd.DataFrame({"date": pd.date_range(start, end)})
        all_days["date"] = pd.to_datetime(all_days["date"])  # 🔧 Ensure datetime64[ns]

        return (
            all_days.merge(df, on="date", how="left")
            .fillna(0)
            .astype({"actual": int, "forecast": int})
        )    
# ─────────── 3. LOAD DATA ------------------------------------------------
    actual_df = get_actual_hours(sel_dept, sel_pos, dt_from, dt_to)
    sched_df  = get_schedule_hours(sel_dept, sel_pos, dt_from, dt_to)
    rooms_df  = get_room_stats("Occupied Rooms", dt_from, dt_to)
    std_df    = get_standard_hours(sel_dept, sel_pos, dt_from, dt_to)

    actual_df["date"] = pd.to_datetime(actual_df["date"])
    sched_df["date"]  = pd.to_datetime(sched_df["date"])
    rooms_df["date"]  = pd.to_datetime(rooms_df["date"])

    # --- merge actual + schedule -------------------------------------------------
    merged = actual_df.merge(sched_df, on="date", how="left").fillna(0)
    merged = merged.merge(std_df, on="date", how="left").fillna({"standard_hours": 0})

    # --- add Occupied-Rooms actuals  --------------------------------------------
    merged = merged.merge(
        rooms_df[["date", "actual"]].rename(columns={"actual": "occ_rooms"}),
        on="date", how="left"
    ).fillna({"occ_rooms": 0})

    # --- calculate standard productivity index (standard hours / occupied rooms)
    merged["standard_prod_idx"] = merged.apply(
        lambda r: r["standard_hours"] / r["occ_rooms"] if r["occ_rooms"] else 0,
        axis=1
    )

    # --- day-level KPIs ----------------------------------------------------------
    merged["actual_fte"] = (merged["hours"] + merged["ot_hours"]) / 8
    merged["sched_fte"]  = merged["sched_hours"] / 8
    merged["prod_idx"]   = merged.apply(
        lambda r: (r["actual_fte"] * 8) / r["occ_rooms"] if r["occ_rooms"] else 0,
        axis=1
    )

    # --- schedule productivity (sched hours / forecast rooms) --------------------
    merged = merged.merge(
        rooms_df[["date", "forecast"]].rename(columns={"forecast": "fc_rooms"}),
        on="date", how="left"
    ).fillna({"fc_rooms": 0})

    merged["sched_prod_idx"] = merged.apply(
        lambda r: r["sched_hours"] / r["fc_rooms"] if r["fc_rooms"] else 0,
        axis=1
    )
    # --- totals / period KPIs ----------------------------------------------------
    tot_act_fte = merged["actual_fte"].sum()
    tot_sch_fte = merged["sched_fte"].sum()
    tot_ot_pct  = (
        merged["ot_hours"].sum() / merged["hours"].sum() * 100
        if merged["hours"].sum() else 0
    )
    prod_index  = (tot_sch_fte / tot_act_fte * 100) if tot_act_fte else 0
    # ─────────── Standard Productivity (from LaborStandard table) ───────────
    std_df = refresh(db.LaborStandard)
    pos_df = refresh(db.Position)
    dept_df = refresh(db.Department)

    std_df = std_df.merge(pos_df.rename(columns={"id": "position_id", "name": "position", "department_id": "dept_id"}), on="position_id", how="left")
    std_df = std_df.merge(dept_df.rename(columns={"id": "dept_id", "name": "dept"}), on="dept_id", how="left")

    std_df = std_df[std_df["metric"] == "Occupied Rooms"]

    if sel_dept != "All":
        std_df = std_df[std_df["dept"] == sel_dept]
    if sel_pos != "All":
        std_df = std_df[std_df["position"] == sel_pos]

    raw_std = std_df["standard"].mean() if not std_df.empty else None
    standard_prod_val = (8 / raw_std) if raw_std else 0
    merged["standard_prod_idx"] = standard_prod_val  # assign same value across chart

    # ─────────── 4. KPI CARDS ------------------------------------------------
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Actual FTEs",  f"{tot_act_fte:.1f}")
    k2.metric("Earned FTEs",  f"{tot_sch_fte:.1f}",
              delta=f"{tot_sch_fte - tot_act_fte:+.1f}")
    k3.metric("Overtime %",   f"{tot_ot_pct:.1f} %")
    k4.metric("Prod Index",   f"{prod_index:.1f}")

    # ─────────── 5. CHARTS ROW ----------------------------------------------
    c_left, c_right = st.columns(2)
    is_single_day   = (dt_from == dt_to)

    # Clean label for x-axis
    merged["x_label"] = merged["date"].dt.strftime("%b %d")

    # ---------- FTE VARIANCE -------------------------------------------------
    with c_left:
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        if is_single_day:
            act_fte = merged["actual_fte"].iloc[0]
            sch_fte = merged["sched_fte"].iloc[0]
            fig.add_trace(go.Indicator(
                mode   = "number+delta",
                value  = act_fte,
                number = {
                    "valueformat": ".2f",
                    "font": {"size": 64, "color": "#1f77b4"},
                    "suffix": (
                        f"<span style='font-size:0.6em; color:#d62728'>"
                        f" / {sch_fte:.2f} Sch</span>"
                    )
                },
                delta  = {
                    "reference": sch_fte,
                    "relative" : False,
                    "position" : "bottom",
                    "increasing": {"color": "green"},
                    "decreasing": {"color": "red"}
                },
                title  = {"text": "<b>Actual FTEs vs Schedule FTEs</b>"},
                domain = {'x': [0, 1], 'y': [0, 1]}
            ))
            fig.update_layout(height=280, margin=dict(t=40, l=0, r=0, b=0))
        else:
            fig.add_trace(go.Scatter(
                x=merged["x_label"],
                y=merged["actual_fte"],
                name="Actual",
                mode="lines+markers",
                marker_color="#1f77b4",
                hovertemplate="Actual&nbsp;FTE: %{y:.2f}<extra></extra>"
            ))
            fig.add_trace(go.Scatter(
                x=merged["x_label"],
                y=merged["sched_fte"],
                name="Schedule",
                mode="lines+markers",
                marker_color="firebrick",
                line=dict(dash="dash"),
                hovertemplate="Sched&nbsp;FTE: %{y:.2f}<extra></extra>"
            ), secondary_y=True)

            fig.update_layout(
                title_text="FTE Variance",
                yaxis_title="FTEs",
                height=280,
                margin=dict(t=40, l=0, r=0, b=0),
                xaxis_type="category",
                hovermode="x unified"              # stacked hover tooltip
            )
        st.plotly_chart(fig, use_container_width=True)
        # ────────────── Totals Summary (FTEs with Variance) ──────────────
        total_actual_fte = merged["actual_fte"].sum()
        total_sched_fte  = merged["sched_fte"].sum()
        variance_fte     = total_sched_fte - total_actual_fte
        variance_arrow   = "▲" if variance_fte > 0 else ("▼" if variance_fte < 0 else "")

        st.markdown(
            f"""
            <div style='text-align: right; font-size: 16px; padding-top: 8px;'>
                <b>Total Actual FTEs:</b> {total_actual_fte:.2f} &nbsp;&nbsp;
                <b>Schedule:</b> {total_sched_fte:.2f} &nbsp;&nbsp;
                <b>Variance:</b> {variance_fte:+.2f} {variance_arrow}
            </div>
            """,
            unsafe_allow_html=True
        )

    # ────────────── Standard Productivity Index (Dashboard) ──────────────
    std_df = refresh(db.LaborStandard)
    pos_df = refresh(db.Position).rename(columns={"id": "position_id", "name": "position", "department_id": "dept_id"})
    dept_df = refresh(db.Department).rename(columns={"id": "dept_id", "name": "dept"})

    std_df = std_df.merge(pos_df, on="position_id", how="left")
    std_df = std_df.merge(dept_df, on="dept_id", how="left")

    std_df = std_df[std_df["metric"] == "Occupied Rooms"]
    if sel_dept != "All":
        std_df = std_df[std_df["dept"] == sel_dept]
    if sel_pos != "All":
        std_df = std_df[std_df["position"] == sel_pos]

    avg_standard = std_df["standard"].mean() if not std_df.empty else None

    total_output = rooms_df["actual"].sum()
    if total_output > 0 and avg_standard:
        standard_prod_val = 8 / avg_standard
    else:
        standard_prod_val = 0.0
    # ---------- PRODUCTIVITY ------------------------------------------------
    with c_right:
        if is_single_day:
            prod_val       = merged["prod_idx"].iloc[0]
            sched_prod_val = standard_prod_val  # from computed standard above

            fig = go.Figure(go.Indicator(
                mode   = "number+delta",
                value  = prod_val,
                number = {
                    "valueformat": ".2f",
                    "font"  : {"size": 64, "color": "#268bd2"},
                    "suffix": (
                        f"<span style='font-size:0.6em; color:#d62728'>"
                        f" / {sched_prod_val:.2f} Std</span>"
                    )
                },
                delta  = {
                    "reference"  : sched_prod_val,
                    "relative"   : False,
                    "position"   : "bottom",
                    "increasing" : {"color": "green"},
                    "decreasing" : {"color": "red"}
                },
                title  = {"text": "<b>Productivity (Hrs / Room)</b>"},
                domain = {'x': [0, 1], 'y': [0, 1]}
            ))
            fig.update_layout(height=280, margin=dict(t=40, l=0, r=0, b=0))

        else:
            fig = go.Figure()

            fig.add_trace(go.Scatter(
                x=merged["x_label"],
                y=merged["prod_idx"],
                mode="lines+markers",
                name="Actual",
                marker_color="#268bd2",
                hovertemplate="Actual&nbsp;Prod: %{y:.2f}<extra></extra>"
            ))

            fig.add_trace(go.Scatter(
                x=merged["x_label"],
                y=[standard_prod_val] * len(merged),
                mode="lines",
                name="Standard",
                marker_color="#d62728",
                line=dict(dash="dash"),
                hovertemplate="Standard&nbsp;Prod: %{y:.2f}<extra></extra>"
            ))

            fig.update_layout(
                title_text="Productivity (Hours / Room)",
                yaxis_title="Hrs / Room",
                height=280,
                margin=dict(t=40, l=0, r=0, b=0),
                xaxis_type="category",
                hovermode="x unified"
            )

        st.plotly_chart(fig, use_container_width=True)
        # ────────────── Totals Summary (Productivity with Variance) ──────────────
        total_hours = merged["hours"].sum() + merged["ot_hours"].sum()
        total_rooms = merged["occ_rooms"].sum()
        actual_prod_total = total_hours / total_rooms if total_rooms else 0
        prod_variance = standard_prod_val - actual_prod_total
        prod_arrow = "▲" if prod_variance > 0 else ("▼" if prod_variance < 0 else "")

        st.markdown(
            f"""
            <div style='text-align: right; font-size: 16px; padding-top: 8px;'>
                <b>Total Actual Prod:</b> {actual_prod_total:.2f} hrs/room &nbsp;&nbsp;
                <b>Standard:</b> {standard_prod_val:.2f} hrs/room &nbsp;&nbsp;
                <b>Variance:</b> {prod_variance:+.2f} {prod_arrow}
            </div>
            """,
            unsafe_allow_html=True
        )
    # ─────────── 7. DAILY ACTUAL vs FORECAST  (pick KPI) -----------------------
    kpi_options = ["Occupied Rooms", "Arrivals", "Covers", "Departures", "Check-outs"]

    sel_kpi = st.selectbox("🛎️ KPI to plot (daily view)", kpi_options, key="kpi_daily")

    # fetch day-level data for the chosen KPI
    kpi_day_df = get_room_stats(sel_kpi, dt_from, dt_to).copy()
    kpi_day_df["x_label"] = kpi_day_df["date"].dt.strftime("%b %d")

    st.markdown(f"#### {sel_kpi} — Actual vs Forecast by Day")

    fig = go.Figure()

    fig.add_trace(go.Bar(
        x = kpi_day_df["x_label"],
        y = kpi_day_df["actual"],
        name = "Actual",
        marker_color = "#1f77b4",
        hovertemplate = "%{x}<br>Actual: %{y}<extra></extra>"
    ))

    fig.add_trace(go.Bar(
        x = kpi_day_df["x_label"],
        y = kpi_day_df["forecast"],
        name = "Forecast",
        marker_color = "#a0a0a0",
        hovertemplate = "%{x}<br>Forecast: %{y}<extra></extra>"
    ))

    fig.update_layout(
        barmode     = "group",
        height      = 300,
        margin      = dict(t=10, l=40, r=40, b=40),
        xaxis_title = "",
        yaxis_title = "Rooms",
        xaxis_type  = "category",
        xaxis_showgrid = False,
        yaxis       = dict(showgrid=True, gridcolor="#e0e0e0"),
        hovermode   = "x unified"
    )

    st.plotly_chart(fig, use_container_width=True)

# ===================================================== EMPLOYEES ==============
elif main_choice == "Employees":
    st.header("👥 Employees")

    tab1, tab2, tab3 = st.tabs(["📋 View / Edit", "➕ Add New", "📆 Schedule Availability"])

    # ---------- 1) VIEW / EDIT -----------------------------------------------
    with tab1:

        # ✅ Show success message if available
        if "emp_delete_msg" in st.session_state:
            st.success(st.session_state.pop("emp_delete_msg"))

        # ▸─ Load employee data ───────────────────────────────────────────────
        emp_df = refresh(db.Employee)

        # ▸─ Filters (top row) ───────────────────────────────────────────────
        st.markdown("#### Filters")
        filter_col1, filter_col2, filter_col3 = st.columns(3)

        with filter_col1:
            depts = sorted(emp_df["department"].dropna().unique())
            sel_dept = st.selectbox("Department", ["(All)"] + depts)

        with filter_col2:
            if sel_dept != "(All)":
                pos_opts = emp_df.loc[
                    emp_df["department"] == sel_dept, "role"
                ].dropna().unique()
                sel_pos = st.selectbox("Position", ["(All)"] + sorted(pos_opts))
            else:
                sel_pos = "(All)"

        with filter_col3:
            emp_opts = ["(Show All)"] + sorted(emp_df["name"].dropna().unique())
            sel_emp = st.selectbox("🔍 Search employee", options=emp_opts, key="emp_selectbox")

        # ▸─ Apply filters ───────────────────────────────────────────────────
        filtered = emp_df.copy()
        if sel_dept != "(All)":
            filtered = filtered[filtered["department"] == sel_dept]
        if sel_pos != "(All)":
            filtered = filtered[filtered["role"] == sel_pos]
        if sel_emp != "(Show All)":
            filtered = filtered[filtered["name"] == sel_emp]

        # ── Deduplicate: keep highest hourly rate per name ──────────────────
        dedup = (
            filtered.sort_values("hourly_rate", ascending=False)
                    .drop_duplicates(subset="name", keep="first")
                    .reset_index(drop=True)
        )

        # ── Add "Type" column (In-House / Contract) ─────────────────────────
        if "emp_type" in dedup.columns:
            dedup["Type"] = np.where(
                dedup["emp_type"].astype(str).str.strip().str.casefold() == "manual",
                "Contract Labor",
                "In-House"
            )
        else:
            dedup["Type"] = "In-House"

        # ── Split name to Last / First / ID ─────────────────────────────────
        parts = dedup["name"].str.extract(
            r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
        )
        for col in ["Last_Name", "First_Name", "ID"]:
            dedup[col] = parts[col]

        # ── Preferred column order ─────────────────────────────────────────
        display_cols = ["Last_Name", "First_Name", "ID", "Type"] + [
            c for c in dedup.columns
            if c not in ["id", "name", "Last_Name", "First_Name", "ID", "Type"]
        ] + ["id", "name"]

        # ── Build editable AgGrid ──────────────────────────────────────────
        gb_emp = GridOptionsBuilder.from_dataframe(dedup[display_cols])
        gb_emp.configure_default_column(editable=True, resizable=True)

        # Header and cell alignment styling
        gb_emp.configure_column("Last_Name", header_name="Last Name", cellStyle={'textAlign': 'left'})
        gb_emp.configure_column("First_Name", header_name="First Name", cellStyle={'textAlign': 'left'})
        gb_emp.configure_column("ID", cellStyle={'textAlign': 'center'})
        gb_emp.configure_column("hourly_rate", cellStyle={'textAlign': 'center'})
        gb_emp.configure_column("department", cellStyle={'textAlign': 'left'})
        gb_emp.configure_column("role", cellStyle={'textAlign': 'left'})
        gb_emp.configure_column("Type", editable=False, cellStyle={'textAlign': 'center'})

        # Hidden columns
        gb_emp.configure_column("id", hide=True)
        gb_emp.configure_column("name", hide=True)

        # Enable selection
        gb_emp.configure_selection("multiple", use_checkbox=True)

        grid_opts = gb_emp.build()

        # Optional: rounded corners style
        st.markdown("""
            <style>
            .ag-theme-streamlit .ag-root-wrapper {
                border-radius: 12px !important;
            }
            </style>
        """, unsafe_allow_html=True)

        grid = AgGrid(
            dedup[display_cols],
            gridOptions=grid_opts,
            theme="streamlit",
            fit_columns_on_grid_load=True,
            domLayout='autoHeight'
        )
        # ── Save button ────────────────────────────────────────────────────
        if st.button("💾 Save edits"):
            for row in grid["data"]:
                session.query(db.Employee).filter_by(id=row["id"]).update(
                    {
                        "name": row["name"],
                        "role": row["role"],
                        "department": row["department"],
                        "hourly_rate": row["hourly_rate"],
                    }
                )
            session.commit()
            st.success("Employee edits saved.")
            st.rerun()

        # ── Delete button ──────────────────────────────────────────────────
        if st.button("🗑️ Delete selected employees"):
            sel_raw = grid.get("selected_rows", [])
            sel_rows = (
                sel_raw.to_dict("records")
                if isinstance(sel_raw, pd.DataFrame) else list(sel_raw)
            )

            if not sel_rows:
                st.warning("No employees selected.")
            else:
                names_to_delete = {
                    r["name"].strip()
                    for r in sel_rows
                    if isinstance(r, dict) and "name" in r and pd.notna(r["name"])
                }
                if names_to_delete:
                    deleted = (
                        session.query(db.Employee)
                               .filter(db.Employee.name.in_(names_to_delete))
                               .delete(synchronize_session=False)
                    )
                    session.commit()
                    st.session_state["emp_delete_msg"] = (
                        f"Deleted {deleted} record(s) for {len(names_to_delete)} employee(s)."
                    )
                    st.rerun()
                else:
                    st.warning("Selected rows contained no valid names.")
    # ---------- 2) ADD NEW ----------------------------------------------------
    with tab2:
        st.subheader("Add Employee Manually")

        # ── persistent banner after a manual add ───────────────────────────
        if "emp_add_msg" in st.session_state:
            st.success(st.session_state.pop("emp_add_msg"))

        # fetch unique Departments & Positions
        depts     = sorted({d.name for d in session.query(db.Department).all()})
        positions = sorted({p.name for p in session.query(db.Position).all()})

        if not depts or not positions:
            st.warning("Create some departments and positions first in the Structure page.")
        else:
            with st.form("emp_form", clear_on_submit=True):
                col1, col2 = st.columns(2)

                # left-hand fields
                with col1:
                    first_name = st.text_input("First Name")
                    last_name  = st.text_input("Last Name")
                    emp_id     = st.text_input("ID #")
                    emp_type   = st.selectbox("Type", ["In-House", "Contract Labor"])

                # right-hand dropdowns + hourly rate
                with col2:
                    dept_sel = st.selectbox("Department", depts)
                    pos_sel  = st.selectbox("Position / Role", positions)
                    rate     = st.number_input("Hourly Rate", min_value=0.0)

                if st.form_submit_button("Add Employee") and first_name and last_name and emp_id:
                    full_name = f"{last_name.strip()}, {first_name.strip()} {emp_id.strip()}"
                    source_val = "manual" if emp_type == "Contract Labor" else "import"

                    session.merge(
                        db.Employee(
                            name        = full_name,
                            role        = pos_sel,
                            department  = dept_sel,
                            hourly_rate = rate,
                            emp_type    = source_val    # ← FIXED
                        )
                    )
                    session.commit()
                    st.session_state["emp_add_msg"] = (
                        f"Employee “{first_name.strip()} {last_name.strip()}” added successfully."
                    )
                    st.rerun()
        # ------------- BULK UPLOAD ------------------------------------------
        st.subheader("📥 Bulk Upload Employees")
        csv_up = st.file_uploader(
            "Upload CSV/XLSX with columns "
            "'Location Name', 'Employee', 'Job Name', 'Base Salary', 'Base Rate'",
            type=["csv", "xls", "xlsx"]
        )
        if csv_up:
            df_up = (pd.read_excel(csv_up)
                     if csv_up.name.endswith((".xls", ".xlsx"))
                     else pd.read_csv(csv_up))

            # read Number as a string so leading zeros survive
      #      pay_df = pd.read_csv(pay_file, dtype={"Number": str})

            required = {"Location Name", "Employee", "Job Name", "Base Rate"}
            if not required.issubset(df_up.columns):
                st.error(f"Missing columns: {required - set(df_up.columns)}")
            else:
                # normalise helper cols
                df_up["loc_norm"] = (
                    df_up["Location Name"].astype(str).str.strip().str.casefold()
                )
                df_up["job_norm"] = (
                    df_up["Job Name"].astype(str).str.strip().str.casefold()
                )

                # ---------- Departments -------------------------------------
                exist_depts_norm = {
                    d.name.strip().casefold()
                    for d in session.query(db.Department).all()
                }
                new_dept_rows = (
                    df_up.loc[~df_up["loc_norm"].isin(exist_depts_norm),
                              ["Location Name", "loc_norm"]]
                    .drop_duplicates("loc_norm")
                )
                for _, row in new_dept_rows.iterrows():
                    session.add(db.Department(name=row["Location Name"].strip()))
                session.commit()

                # map norm dept → id
                dept_map = {
                    d.name.strip().casefold(): d.id
                    for d in session.query(db.Department).all()
                }

                # ---------- Positions ---------------------------------------
                exist_pos_keys = {
                    (p.name.strip().casefold(), p.department_id)
                    for p in session.query(db.Position).all()
                }
                new_pos_added = 0
                for _, row in df_up.drop_duplicates(
                    subset=["job_norm", "loc_norm"]
                ).iterrows():
                    dept_id = dept_map.get(row["loc_norm"])
                    key = (row["job_norm"], dept_id)
                    if key not in exist_pos_keys and dept_id is not None:
                        session.add(
                            db.Position(
                                name=row["Job Name"].strip(),
                                department_id=dept_id
                            )
                        )
                        new_pos_added += 1
                session.commit()

                # ---------- Employees ---------------------------------------
                added = 0
                updated = 0
                skipped = 0

                for _, r in df_up.iterrows():
                    full_name = str(r["Employee"]).strip()
                    dept_norm = str(r["Location Name"]).strip().casefold()
                    job_norm  = str(r["Job Name"]).strip().casefold()
                    rate_val  = float(r["Base Rate"])

                    dept_id = dept_map.get(dept_norm)
                    if dept_id is None:
                        continue

                    # Find position object by name + department
                    pos = (
                        session.query(db.Position)
                        .filter(
                            db.Position.name.ilike(r["Job Name"].strip()),
                            db.Position.department_id == dept_id
                        )
                        .first()
                    )
                    if not pos:
                        continue

                    # Check if a manual employee (Contract Labor) exists
                    manual_exists = (
                        session.query(db.Employee)
                        .filter_by(name=full_name, emp_type="manual")
                        .first()
                    )
                    if manual_exists:
                        skipped += 1
                        continue

                    # Now add/update only IMPORT employees
                    emp = (
                        session.query(db.Employee)
                        .filter_by(name=full_name, emp_type="import")
                        .first()
                    )

                    if emp:
                        emp.role        = pos.name
                        emp.department  = r["Location Name"].strip()
                        emp.hourly_rate = rate_val
                        updated += 1
                    else:
                        session.add(
                            db.Employee(
                                name        = full_name,
                                role        = pos.name,
                                department  = r["Location Name"].strip(),
                                hourly_rate = rate_val,
                                emp_type    = "import"
                            )
                        )
                        added += 1

                session.commit()
                st.success(f"✅ Employees processed: {added} added · {updated} updated · {skipped} skipped (manual)")
                # ---------- Employees (add) ----------------------------------
                exist_emp = {
                    e.name.strip().casefold()
                    for e in session.query(db.Employee).all()
                }
                new_emp_added = 0
                for _, row in df_up.iterrows():
                    emp_name = str(row["Employee"]).strip()
                    if emp_name.casefold() in exist_emp:
                        continue
                    session.add(
                        db.Employee(
                            name=emp_name,
                            role=row["Job Name"].strip(),
                            department=row["Location Name"].strip(),
                            hourly_rate=float(row["Base Rate"]) if pd.notna(row["Base Rate"]) else 0.0,
                        )
                    )
                    new_emp_added += 1
                session.commit()

                # ---------- DELETE Employees not in upload -------------------
                names_in_upload_norm = {
                    str(n).strip().lower() for n in df_up["Employee"].dropna()
                }

                deleted = (
                    session.query(db.Employee)
                           .filter(
                               db.Employee.emp_type == "import",                # ← keep Contract Labor
                               ~func.lower(db.Employee.name).in_(names_in_upload_norm)
                           )
                           .delete(synchronize_session=False)
                )
                session.commit()

                # ---------- Feedback banner ----------------------------------
                st.success(
                    f"Imported {len(new_dept_rows)} new departments, "
                    f"{new_pos_added} new positions, {new_emp_added} new employees. "
                    f"Removed {deleted} obsolete In-House employee record(s)."
                )

    # ---------- 3) SCHEDULE AVAILABILITY --------------------------------------
    with tab3:
        st.subheader("Weekly Schedule Availability")

        emp_df = refresh(db.Employee)

        # ── All existing availability rows (dummy week only) ───────────────
        dummy_dates = list(WK_TO_DATE.values())      # 2000-01-03 … 2000-01-09
        sched_rows = session.query(db.Schedule).filter(
            db.Schedule.day.in_(dummy_dates)         # ← filter here
        ).all()

        # ── Filter sidebar ────────────────────────────────────────────────
        col_left, col_main = st.columns([1, 3])

        with col_left:
            st.markdown("#### Filter")
            depts = sorted(emp_df["department"].dropna().unique())
            sel_dept = st.selectbox("Department", ["(All)"] + depts, key="view_dept")

            if sel_dept != "(All)":
                pos_opts = emp_df[emp_df["department"] == sel_dept]["role"].dropna().unique()
                sel_pos  = st.selectbox("Position", ["(All)"] + sorted(pos_opts))
            else:
                sel_pos = "(All)"

            emp_opts = emp_df["name"].dropna().unique()
            sel_emp = st.selectbox("Employee", ["(All)"] + sorted(emp_opts))

        # ── Apply filters ─────────────────────────────────────────────────
        filtered = emp_df.copy()
        if sel_dept != "(All)":
            filtered = filtered[filtered["department"] == sel_dept]
        if sel_pos != "(All)":
            filtered = filtered[filtered["role"] == sel_pos]
        if sel_emp != "(All)":
            filtered = filtered[filtered["name"] == sel_emp]

        filtered = filtered.drop_duplicates(subset=["name"])  # one row per employee

        # ── Build editable availability table ─────────────────────────────
        import re
        from datetime import date, datetime
        days          = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        shift_options = ["MORNING", "AFTERNOON", "EVENING", "OPEN", "OFF"]

        data = []
        pattern = re.compile(r"^\s*(?P<last>[^,]+),\s*(?P<first>[^\d]+?)\s+(?P<id>\d+)")

        for _, row in filtered.iterrows():
            emp_id   = row["id"]
            raw_name = row["name"]

            m = pattern.match(raw_name)
            if m:
                last_name  = m.group("last").strip()
                first_name = m.group("first").strip()
                id_str     = m.group("id").strip()
            else:
                last_name, first_name, id_str = raw_name, "", ""

            # Existing availability (dummy-week rows only)
            shifts = {}
            for s in sched_rows:
                if s.emp_id == emp_id:
                    lbl = s.day.strftime("%a")  # 'Mon', 'Tue', …
                    shifts[lbl] = s.shift_type

            row_data = {
                "ID":         id_str,
                "First Name": first_name,
                "Last Name":  last_name,
                "emp_id":     emp_id
            }
            for d in days:
                row_data[d] = shifts.get(d, "OPEN")
            data.append(row_data)

        schedule_edit_df = pd.DataFrame(data)

        # ── Configure AgGrid ─────────────────────────────────────────────
        gb = GridOptionsBuilder.from_dataframe(schedule_edit_df)
        gb.configure_default_column(editable=True)

        gb.configure_column("ID",         width=160, editable=False)
        gb.configure_column("First Name", width=160, editable=False)
        gb.configure_column("Last Name",  width=160, editable=False)

        for d in days:
            gb.configure_column(
                d,
                cellEditor="agSelectCellEditor",
                cellEditorParams={"values": shift_options},
                width=160
            )

        gb.configure_column("emp_id", hide=True)
        gb.configure_selection("multiple", use_checkbox=False)

        grid = AgGrid(
            schedule_edit_df,
            gridOptions=gb.build(),
            height=400,
            theme="balham"
        )

        # ── Save button ──────────────────────────────────────────────────
        if st.button("💾 Save Schedule"):
            edited = pd.DataFrame(grid["data"])

            # Delete existing dummy-week rows for these employees
            session.query(db.Schedule).filter(
                db.Schedule.emp_id.in_(edited["emp_id"].tolist()),
                db.Schedule.day.in_(dummy_dates)
            ).delete(synchronize_session=False)

            # Re-insert availability rows (always dummy-week dates)
            new_rows = []
            for _, r in edited.iterrows():
                emp = int(r["emp_id"])
                for lbl in days:
                    new_rows.append(
                        db.Schedule(
                            emp_id     = emp,
                            day        = WK_TO_DATE[lbl],          # already a date obj
                            shift_type = str(r[lbl]).strip().upper()
                        )
                    )

            session.bulk_save_objects(new_rows)
            session.commit()
            st.success("✅ Weekly availability saved.")
# =============================================== LABOR ▸ STRUCTURE ============
elif main_choice == "Labor ▸ Structure":
    st.header("🏗️ Labor Structure")

    tab1, tab2, tab3, tab4 = st.tabs(["📋 View / Edit", "➕ Add New", "🕒 Labor Shifts", "📏 Labor Standards"])

    # ---------- 1) VIEW / EDIT -----------------------------------------------
    with tab1:

        #–– PERSISTENT MESSAGE ───────────────────────────────────────────────
        if "del_msg" in st.session_state:
            st.success(st.session_state.pop("del_msg"))

        st.markdown("""
            <style>
            .ag-theme-streamlit .ag-root-wrapper {
                border-radius: 12px !important;
            }
            </style>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns([1, 2])

        with col1:
            st.subheader("Departments")
            dept_df = refresh(db.Department)
            gb_dept = GridOptionsBuilder.from_dataframe(dept_df)
            gb_dept.configure_default_column(editable=True, resizable=True)
            gb_dept.configure_selection("multiple", use_checkbox=True)

            dept_grid = AgGrid(
                dept_df,
                gridOptions=gb_dept.build(),
                theme="streamlit",
                fit_columns_on_grid_load=True,
                height=350
            )
            sel_depts = dept_grid.get("selected_rows", [])

        with col2:
            st.subheader("Positions")
            pos_df = refresh(db.Position).merge(
                dept_df, left_on="department_id", right_on="id",
                suffixes=("", "_dept"))
            gb_pos = GridOptionsBuilder.from_dataframe(pos_df)
            gb_pos.configure_default_column(editable=True, resizable=True)
            gb_pos.configure_selection("multiple", use_checkbox=True)

            pos_grid = AgGrid(
                pos_df,
                gridOptions=gb_pos.build(),
                theme="streamlit",
                fit_columns_on_grid_load=True,
                height=350
            )
            sel_pos = pos_grid.get("selected_rows", [])

        col_a, col_b = st.columns(2)
        if col_a.button("💾 Save edits"):
            # save department edits
            for row in dept_grid["data"]:
                session.query(db.Department).filter_by(id=row["id"]).update({"name": row["name"]})
            # save position edits
            for row in pos_grid["data"]:
                session.query(db.Position).filter_by(id=row["id"]).update({"name": row["name"]})
            session.commit()
            st.success("Changes saved.")
            st.rerun()

        if col_b.button("🗑️ Delete selected"):
            # ---- helper to ensure we always iterate over a list ------------
            def _rows(sel):
                if sel is None:
                    return []
                if isinstance(sel, list):
                    return sel
                if isinstance(sel, pd.DataFrame):
                    return sel.to_dict("records")
                return []

            pos_rows  = _rows(sel_pos)
            dept_rows = _rows(sel_depts)

            # ---- delete positions first -----------------------------------
            for r in pos_rows:
                rec_id = r["id"] if isinstance(r, dict) and "id" in r else r
                try:
                    rec_id = int(rec_id)
                    session.query(db.Position).filter_by(id=rec_id).delete()
                except Exception:
                    pass  # skip invalid ids

            # ---- then delete departments (skip if referenced) -------------
            for r in dept_rows:
                rec_id  = r["id"] if isinstance(r, dict) and "id" in r else r
                r_name  = r.get("name", str(r)) if isinstance(r, dict) else str(r)
                try:
                    rec_id = int(rec_id)
                    session.query(db.Department).filter_by(id=rec_id).delete()
                except Exception:
                    st.warning(f"Dept '{r_name}' not deleted (still referenced).")

            session.commit()
            #–– SAVE BANNER ─────────────────────────────────────────────────
            st.session_state["del_msg"] = (
                f"Deleted {len(pos_rows)} positions and {len(dept_rows)} departments."
            )
            st.rerun()

    # ---------- 2) ADD NEW ----------------------------------------------------
    with tab2:
        st.subheader("Add Department")
        with st.form("dept_form", clear_on_submit=True):
            new_dept = st.text_input("Department name")
            if st.form_submit_button("Add Department") and new_dept:
                session.merge(db.Department(name=new_dept))
                session.commit()
                st.success("Department added!")
                st.rerun()

        st.subheader("Add Position")
        depts = refresh(db.Department)
        if depts.empty:
            st.warning("Create a department first.")
        else:
            with st.form("pos_form", clear_on_submit=True):
                pos_name = st.text_input("Position name")
                dept_sel = st.selectbox("Department", depts["name"])
                if st.form_submit_button("Add Position") and pos_name:
                    dept_id = int(depts.set_index("name").loc[dept_sel, "id"])
                    session.merge(db.Position(name=pos_name, department_id=dept_id))
                    session.commit()
                    st.success("Position added!")
                    st.rerun()

        # Bulk upload (structure)  – unchanged
        st.subheader("📥 Bulk Upload (Location Name → Department, Job Name → Position)")
        csv_file = st.file_uploader(
            "Upload CSV/XLSX with columns 'Location Name' and 'Job Name'",
            type=["csv", "xls", "xlsx"],
            key="struct_upl"
        )
        if csv_file:
            df_up = (pd.read_excel(csv_file)
                     if csv_file.name.endswith((".xls", ".xlsx"))
                     else pd.read_csv(csv_file))

            if not {"Location Name", "Job Name"}.issubset(df_up.columns):
                st.error("Required columns not found.")
            else:
                df_up["loc_norm"] = df_up["Location Name"].astype(str).str.strip().str.casefold()
                df_up["job_norm"] = df_up["Job Name"].astype(str).str.strip().str.casefold()

                exist_depts_norm = {
                    d.name.strip().casefold()
                    for d in session.query(db.Department).all()
                }
                new_dept_rows = (
                    df_up.loc[~df_up["loc_norm"].isin(exist_depts_norm),
                              ["Location Name", "loc_norm"]]
                    .drop_duplicates("loc_norm")
                )
                for _, row in new_dept_rows.iterrows():
                    session.add(db.Department(name=row["Location Name"].strip()))
                session.commit()

                dept_map = {
                    d.name.strip().casefold(): d.id
                    for d in session.query(db.Department).all()
                }

                exist_pos_keys = {
                    (p.name.strip().casefold(), p.department_id)
                    for p in session.query(db.Position).all()
                }
                added = 0
                for _, row in (
                    df_up[["Job Name", "job_norm", "loc_norm"]]
                    .dropna()
                    .drop_duplicates()
                ).iterrows():
                    dept_id = dept_map.get(row["loc_norm"])
                    key = (row["job_norm"], dept_id)
                    if key not in exist_pos_keys:
                        session.add(
                            db.Position(
                                name=row["Job Name"].strip(),
                                department_id=dept_id
                            )
                        )
                        added += 1
                session.commit()
                st.success(
                    f"Imported {len(new_dept_rows)} new departments and {added} new positions."
                )

    # ---------- 3) LABOR SHIFTS -------------------------------------------
    with tab3:
        st.subheader("Define Shift Times by Position")

        # 1. Fetch departments / positions
        depts = sorted({d.name for d in session.query(db.Department).all()})
        dept_sel = st.selectbox("Department", ["(All)"] + depts)

        if dept_sel != "(All)":
            dept_positions = (
                session.query(db.Position)
                       .join(db.Department, db.Position.department_id == db.Department.id)
                       .filter(db.Department.name == dept_sel)
                       .order_by(db.Position.name)
                       .all()
            )
            pos_names = [p.name for p in dept_positions]
            sel_pos = st.selectbox("Position", ["(All)"] + pos_names)
        else:
            sel_pos = "(All)"

        # 👉  Save button — at the top
        save_clicked = st.button("💾 Save All Shift Times", key="save_shifts")

        # 2. Filter positions
        pos_query = session.query(db.Position)
        if dept_sel != "(All)":
            dept_id = session.query(db.Department).filter_by(name=dept_sel).first().id
            pos_query = pos_query.filter_by(department_id=dept_id)
        if sel_pos != "(All)":
            pos_query = pos_query.filter_by(name=sel_pos)

        positions = pos_query.all()
        shift_periods = ["Morning", "Afternoon", "Evening"]

        # 3. UI to edit shifts
        for pos in positions:
            st.markdown(f"#### {pos.name}")

            for period in shift_periods:
                existing = (
                    session.query(db.ShiftTime)
                    .filter_by(position_id=pos.id, period=period)
                    .order_by(db.ShiftTime.id)
                    .all()
                )

                with st.expander(f"🕒 {period} shifts for {pos.name}", expanded=bool(existing)):
                    existing_cnt = len(existing)
                    rows = st.number_input(
                        f"# of {period} shifts",
                        min_value=0, max_value=30,
                        value=existing_cnt,
                        key=f"{pos.id}_{period}_count"
                    )

                    for i in range(rows):
                        c1, c2 = st.columns(2)
                        def_time = lambda obj, field: getattr(obj, field) if obj else None
                        start = c1.time_input(
                            "Start",
                            value=def_time(existing[i] if i < existing_cnt else None, "start"),
                            step=1800,
                            key=f"{pos.id}_{period}_start_{i}"
                        )
                        end = c2.time_input(
                            "End",
                            value=def_time(existing[i] if i < existing_cnt else None, "end"),
                            step=1800,
                            key=f"{pos.id}_{period}_end_{i}"
                        )

        # 4. Persist when Save pressed
        if save_clicked:
            # clear only the shifts for displayed positions
            session.query(db.ShiftTime).filter(
                db.ShiftTime.position_id.in_([p.id for p in positions])
            ).delete(synchronize_session=False)

            for pos in positions:
                for period in shift_periods:
                    rows = st.session_state.get(f"{pos.id}_{period}_count", 0)
                    for i in range(rows):
                        start = st.session_state.get(f"{pos.id}_{period}_start_{i}")
                        end = st.session_state.get(f"{pos.id}_{period}_end_{i}")
                        if start and end:
                            session.add(
                                db.ShiftTime(
                                    position_id=pos.id,
                                    period=period,
                                    start=start,
                                    end=end
                                )
                            )
            session.commit()
            st.success("✅ Shifts saved successfully.")

    # ---------- 4) LABOR STANDARDS ------------------------------------------
    with tab4:
        available_metrics = [
            "Occupied Rooms",
            "Arrivals",
            "Covers",
            "Departures",
            "Check-outs",
        ]

        emp_df = refresh(db.Employee)
        depts = emp_df["department"].dropna().unique().tolist()

        col_dept, col_pos = st.columns(2)
        with col_dept:
            sel_dept = st.selectbox("Department", ["(Select)"] + sorted(depts))

        with col_pos:
            if sel_dept != "(Select)":
                pos_opts = (
                    emp_df.loc[emp_df["department"] == sel_dept, "role"]
                    .dropna()
                    .unique()
                    .tolist()
                )
                sel_pos = st.selectbox("Position", ["(Select)"] + sorted(pos_opts))
            else:
                sel_pos = "(Select)"

        if sel_pos == "(Select)":
            st.info("Select a department **and** position to view labor standards.")
            st.stop()

        pos_rec = (
            session.query(db.Position)
            .filter(func.lower(db.Position.name) == sel_pos.lower())
            .first()
        )
        if not pos_rec:
            st.error("Position not found in database.")
            st.stop()
        pos_id = pos_rec.id

        std_df = load_labor_standards(pos_id)
        if std_df.empty:
            std_df = pd.DataFrame([{"Metric": "", "Standard": None}])
        else:
            std_df = std_df[["Metric", "Standard"]].copy()
        std_df["Standard"] = pd.to_numeric(std_df["Standard"], errors="coerce")

        st.markdown("#### Labor Standards for This Position")

        edited_df = st.data_editor(
            std_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "Metric": st.column_config.SelectboxColumn(
                    "Metric", options=available_metrics
                ),
                "Standard": st.column_config.NumberColumn(
                    "Standard per FTE", format="%.2f"
                ),
            },
            key=f"labor_std_editor_{pos_id}"  # ensures reset when switching position
        )

        if st.button("💾 Save Standards"):
            edited_df = edited_df[
                edited_df["Metric"].notna() & edited_df["Standard"].notna()
            ]
            edited_df["Standard"] = edited_df["Standard"].astype(float)
            edited_df["Unit"] = edited_df["Metric"].astype(str) + " per FTE"
            save_labor_standards(pos_id, edited_df)
            st.success("Labor standards saved.")
# =========================================== LABOR ▸ ACTUAL HOURS =============
elif main_choice == "Labor ▸ Actual Hours":
      import os, re, pandas as pd
      from datetime import date, timedelta
      from dateutil.relativedelta import relativedelta, MO
      from sqlalchemy import or_, func
      from st_aggrid import AgGrid, GridOptionsBuilder

      # ────────────────────────── PAGE HEADER ───────────────────────────────
      st.header("⏱️ Actual Hours")

      latest_date = (
            session.query(func.max(db.Actual.date))
                   .filter(or_(db.Actual.hours    != 0,
                               db.Actual.ot_hours != 0,
                               db.Actual.reg_pay  != 0,
                               db.Actual.ot_pay   != 0))
                   .scalar()
      )
      if latest_date:
            st.info(f"📅 Latest actual hours with activity: **{latest_date:%B %d, %Y}**")
      else:
            st.info("📅 No non‑zero actual‑hours data found yet.")

      # ───────────────────────── TAB BAR ────────────────────────────────────
      tab_import, tab_emp = st.tabs(["Import / Enter", "Actual Hours"])

      # =====================================================================
      # TAB A — IMPORT / ENTER (Payroll CSV → db.Actual, Contract editor)
      # =====================================================================
      with tab_import:

            st.subheader("📥 Import Payroll CSV (auto‑map to Manual)")
            pay_file = st.file_uploader(
                  "Drop payroll CSV here",
                  type=["csv"],
                  key="payroll_csv"
            )

            def norm_code(code: str) -> str:
                  return code.strip().casefold().replace(" ", "").replace(".", "")

            if pay_file is not None:

                  pay_df = pd.read_csv(pay_file)

                  # ── column aliases --------------------------------------------------
                  COL_POS   = "Job"
                  COL_DATE  = "Business Date"
                  COL_HOURS = "Hours"
                  COL_AMT   = "Pay Amount"
                  COL_CODE  = "Pay Category"
                  COL_ID    = "Number"

                  pay_df[COL_DATE] = pd.to_datetime(pay_df[COL_DATE]).dt.date

                  pos_lookup = {p.name.strip(): p.id for p in session.query(db.Position).all()}

                  RAW = {
                        "Reg":      ("Reg Hours", "Reg Pay"),
                        "Hol1.0":   ("Reg Hours", "Reg Pay"),
                        "Prem":     ("Reg Hours", "Reg Pay"),
                        "PTO":      ("Reg Hours", "Reg Pay"),
                        "Vacation": ("Reg Hours", "Reg Pay"),
                        "OT 1.5":   ("OT Hours",  "OT Pay"),
                        "OT1.5":    ("OT Hours",  "OT Pay"),
                  }
                  CODE_TO_METRIC = {norm_code(k): v for k, v in RAW.items()}

                  added_rows = skipped_pos = skipped_code = 0
                  examples_pos, examples_code = [], []

                  with session.no_autoflush:
                        overwrite_keys = {
                              (pos_lookup.get(str(r[COL_POS]).strip()), r[COL_DATE])
                              for _, r in pay_df.iterrows()
                              if pos_lookup.get(str(r[COL_POS]).strip())
                        }
                        for pos_id, biz_date in overwrite_keys:
                              session.query(db.Actual).filter_by(
                                    position_id=pos_id,
                                    date=biz_date,
                                    source="manual"
                              ).delete()

                  for _, r in pay_df.iterrows():
                        pos_key = str(r[COL_POS]).strip()
                        pos_id  = pos_lookup.get(pos_key)
                        if not pos_id:
                              skipped_pos += 1
                              if len(examples_pos) < 5:
                                    examples_pos.append(pos_key)
                              continue

                        paycode_norm = norm_code(str(r[COL_CODE]))
                        pair = CODE_TO_METRIC.get(paycode_norm)
                        if not pair:
                              skipped_code += 1
                              if len(examples_code) < 5:
                                    examples_code.append(r[COL_CODE])
                              continue

                        metric_hrs, metric_pay = pair
                        biz_date   = r[COL_DATE]
                        hrs_val    = r[COL_HOURS]
                        pay_val    = r[COL_AMT]
                        emp_id_val = str(r[COL_ID]).strip().removesuffix(".0")

                        rec = db.Actual(
                              emp_id      = emp_id_val,
                              position_id = pos_id,
                              date        = biz_date,
                              hours       = hrs_val if metric_hrs == "Reg Hours" else 0,
                              ot_hours    = hrs_val if metric_hrs == "OT Hours" else 0,
                              reg_pay     = pay_val if metric_pay == "Reg Pay"   else 0,
                              ot_pay      = pay_val if metric_pay == "OT Pay"    else 0,
                              source      = "manual"
                        )
                        session.add(rec)
                        added_rows += 1

                  session.commit()

                  st.success(
                        f"✅ Imported {added_rows} rows · {skipped_pos} unknown positions · {skipped_code} unknown pay‑codes"
                  )
                  if examples_pos:
                        st.warning(f"Unmatched position examples: {examples_pos}")
                  if examples_code:
                        st.warning(f"Unmatched pay‑code examples: {examples_code}")

                  # ── 3) de‑duplicate cached uploads so totals don’t double ——
                  key_cols = [COL_ID, COL_DATE, COL_POS, COL_CODE, COL_HOURS, COL_AMT]

                  if "payroll_cache" in st.session_state and not st.session_state["payroll_cache"].empty:
                        combined = pd.concat(
                              [st.session_state["payroll_cache"], pay_df],
                              ignore_index=True
                        )
                        combined.sort_values(key_cols, inplace=True)
                        combined.drop_duplicates(subset=key_cols, keep="last", inplace=True)
                        st.session_state["payroll_cache"] = combined.reset_index(drop=True)
                  else:
                        st.session_state["payroll_cache"] = pay_df.copy()

            # ────────── POSITION FILTER + WEEK NAVIGATOR + CONTRACT EDITOR ──────────
            st.subheader("Filter")

            dept_df = refresh(db.Department).rename(columns={"name": "dept"})
            pos_df  = (
                  refresh(db.Position)
                  .merge(dept_df, left_on="department_id", right_on="id")
                  [["id_x", "name", "dept"]]
                  .rename(columns={"id_x": "id"})
            )

            dept_opts = ["(All)"] + sorted(dept_df["dept"].dropna().unique())
            f1, f2 = st.columns(2)
            with f1:
                  sel_dept = st.selectbox("Department", dept_opts)
            if sel_dept != "(All)":
                  pos_opts = ["(All)"] + sorted(
                        pos_df.loc[pos_df["dept"] == sel_dept, "name"].dropna().unique()
                  )
            else:
                  pos_opts = ["(All)"]
            with f2:
                  sel_pos = st.selectbox("Position", pos_opts)

            if sel_pos != "(All)":
                  sel_pos_id = int(
                        pos_df.loc[(pos_df["dept"] == sel_dept) &
                                   (pos_df["name"] == sel_pos), "id"].values[0]
                  )
                  sel_path = f"{sel_dept}/{sel_pos}"
            else:
                  sel_pos_id, sel_path = None, None

            if "week_start" not in st.session_state:
                  st.session_state.week_start = date.today() + relativedelta(weekday=MO(-1))

            cprev, crange, cnext = st.columns([1, 3, 1])
            if cprev.button("◀", key="week_prev"):
                  st.session_state.week_start -= timedelta(days=7)
            if cnext.button("▶", key="week_next"):
                  st.session_state.week_start += timedelta(days=7)

            week_start = st.session_state.week_start
            week_end   = week_start + timedelta(days=6)
            days       = [week_start + timedelta(d) for d in range(7)]
            fmt = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
            crange.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

            if sel_pos_id is None:
                  st.info("Select a Department and Position to view or edit hours.")
            else:
                  st.markdown(f"#### {sel_path}")

                  act_rows = (session.query(db.Actual)
                                      .filter(db.Actual.position_id == sel_pos_id,
                                              db.Actual.date.between(week_start, week_end))
                                      .all())
                  df = pd.DataFrame([r.__dict__ for r in act_rows])

                  idx = pd.MultiIndex.from_product(
                        [["Manual", "Contract", "Total"],
                         ["Reg Hours", "OT Hours", "Reg Pay", "OT Pay"]],
                        names=["Block", "Metric"]
                  )
                  tbl = pd.DataFrame(index=idx,
                                     columns=[d.strftime(fmt) for d in days]).fillna(0.0)

                  def _add(block, metric, d, val):
                        tbl.loc[(block, metric), d.strftime(fmt)] += float(val or 0)

                  if not df.empty:
                        for _, row in df.iterrows():
                              block = row["source"].capitalize()
                              _add(block, "Reg Hours", row["date"], row["hours"])
                              _add(block, "OT Hours",  row["date"], row["ot_hours"])
                              _add(block, "Reg Pay",   row["date"], row["reg_pay"])
                              _add(block, "OT Pay",    row["date"], row["ot_pay"])

                  for d in days:
                        col = d.strftime(fmt)
                        for m in ["Reg Hours", "OT Hours", "Reg Pay", "OT Pay"]:
                              tbl.loc[("Total", m), col] = (
                                    tbl.loc[("Manual", m), col] +
                                    tbl.loc[("Contract", m), col]
                              )

                  full_df = tbl.reset_index()
                  st.markdown("""
                        <style>
                        .ag-theme-streamlit .ag-root-wrapper {
                              border-radius: 12px !important;
                        }
                        </style>
                  """, unsafe_allow_html=True)

                  gb = GridOptionsBuilder.from_dataframe(full_df)
                  gb.configure_default_column(resizable=True, flex=1, minWidth=100)
                  gb.configure_default_column(editable=True, type=["numericColumn"], precision=2)
                  gb.configure_column("Block", editable=False, pinned="left")
                  gb.configure_column("Metric", editable=False, pinned="left")
                  gb.configure_grid_options(suppressHorizontalScroll=False)
                  gb.configure_grid_options(forceFitColumns=True)

                  grid = AgGrid(
                        full_df,
                        gridOptions=gb.build(),
                        theme="streamlit",
                        fit_columns_on_grid_load=True,
                        allow_unsafe_jscode=True,
                        domLayout='autoHeight',
                        key=f"grid_{sel_pos_id}_{week_start}"
                  )
                  if st.button("💾 Save Contract entries"):
                        edited_df   = pd.DataFrame(grid["data"])
                        contract_df = edited_df[edited_df["Block"] == "Contract"].set_index("Metric")

                        affected_keys = set()
                        for d in days:
                              col = d.strftime(fmt)
                              session.query(db.Actual).filter(
                                    db.Actual.position_id == sel_pos_id,
                                    db.Actual.date        == d,
                                    db.Actual.source      == "contract"
                              ).delete(synchronize_session=False)

                              for metric in ["Reg Hours", "OT Hours", "Reg Pay", "OT Pay"]:
                                    val = float(contract_df.at[metric, col])
                                    if val == 0 or pd.isna(val):
                                          continue
                                    session.add(db.Actual(
                                          emp_id      = None,
                                          position_id = sel_pos_id,
                                          date        = d,
                                          hours       = val if metric == "Reg Hours" else 0,
                                          ot_hours    = val if metric == "OT Hours" else 0,
                                          reg_pay     = val if metric == "Reg Pay"  else 0,
                                          ot_pay      = val if metric == "OT Pay"   else 0,
                                          source      = "contract"
                                    ))
                              affected_keys.add((sel_pos_id, d))

                        refresh_totals(session, affected_keys)
                        session.commit()
                        st.success("✅ Contract data saved.")

      with tab_emp:

            st.subheader("👥 Employee Actual Hours")

            # ── 1) Filters ---------------------------------------------------
            dept_df = refresh(db.Department).rename(columns={"name": "dept"})
            pos_df  = (
                  refresh(db.Position)
                  .merge(dept_df, left_on="department_id", right_on="id")
                  [["id_x", "name", "dept"]]
                  .rename(columns={"id_x": "id"})
            )

            dept_opts = ["(All)"] + sorted(dept_df["dept"].dropna().unique())
            col1, col2 = st.columns(2)
            with col1:
                  sel_dept = st.selectbox("Department", dept_opts, key="emp_dept")
            if sel_dept != "(All)":
                  pos_opts = ["(All)"] + sorted(
                        pos_df.loc[pos_df["dept"] == sel_dept, "name"].dropna().unique()
                  )
            else:
                  pos_opts = ["(All)"]
            with col2:
                  sel_pos = st.selectbox("Position", pos_opts, key="emp_pos")

            # ── 2) Week navigator -------------------------------------------
            if "emp_week_start" not in st.session_state:
                  st.session_state.emp_week_start = date.today() + relativedelta(weekday=MO(-1))

            col_prev, col_range, col_next = st.columns([1, 3, 1])
            if col_prev.button("◀", key="emp_prev"):
                  st.session_state.emp_week_start -= timedelta(days=7)
            if col_next.button("▶", key="emp_next"):
                  st.session_state.emp_week_start += timedelta(days=7)

            week_start = st.session_state.emp_week_start
            week_end   = week_start + timedelta(days=6)
            days       = [week_start + timedelta(d) for d in range(7)]
            fmt        = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
            fmt_cols   = [d.strftime(fmt) for d in days]
            col_range.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

            # ── 3) Pull Actual rows from DB (week-range + filters) ----------
            q = (
                  session.query(
                        db.Actual.emp_id.label("Number"),
                        db.Actual.date.label("Business Date"),
                        (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
                        db.Position.name.label("Position"),
                        db.Department.name.label("Department")
                  )
                  .join(db.Position,   db.Actual.position_id == db.Position.id)
                  .join(db.Department, db.Position.department_id == db.Department.id)
                  .filter(db.Actual.date.between(week_start, week_end))
                  .filter(or_(db.Actual.hours != 0, db.Actual.ot_hours != 0))
            )

            if sel_dept != "(All)":
                  q = q.filter(db.Department.name == sel_dept)
            if sel_pos != "(All)":
                  q = q.filter(db.Position.name == sel_pos)

            raw = pd.DataFrame(q.all(),
                               columns=["Number", "Business Date", "Hours", "Position", "Department"])

            # Allow blank emp_id (None) and still show hours
            raw["Number"] = raw["Number"].astype(str).str.strip().fillna("")

            raw["Business Date"] = pd.to_datetime(raw["Business Date"]).dt.date

            if raw.empty:
                  st.warning("No rows match the current filters / week.")
                  st.stop()

            # ── 4) Parse employee names -------------------------------------
            emp_df = refresh(db.Employee).copy()
            parts = emp_df["name"].astype(str).str.extract(
                  r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
            )
            emp_df["ID"]         = parts["ID"].fillna("").astype(str).str.strip().str.zfill(5)
            emp_df["First Name"] = parts["First_Name"].str.strip()
            emp_df["Last Name"]  = parts["Last_Name"].str.strip()

            raw = pd.DataFrame(q.all())

            # Drop records with no employee ID (emp_id is None)
            raw = raw[raw["Number"].notna()]

            # Normalize 'Number' to match the format of emp_df["ID"]
            raw["Number"] = (
                  pd.to_numeric(raw["Number"], errors="coerce")
                    .fillna(0)
                    .astype(int)
                    .astype(str)
                    .str.zfill(5)
            )

            # ── 5) Pivot table by day ---------------------------------------
            pivot = (
                  raw.pivot_table(index="Number",
                                  columns="Business Date",
                                  values="Hours",
                                  aggfunc="sum",
                                  fill_value=0)
                     .reset_index()
            )

            # ── Fix ID parsing and cleanup ----------------------------------
            pivot = pivot[pd.to_numeric(pivot["Number"], errors="coerce").notna()]
            pivot["ID"] = (
                  pivot["Number"]
                        .astype(str)
                        .str.strip()
                        .str.replace(r"\.0$", "", regex=True)
            )
            pivot.drop(columns="Number", inplace=True)

            # create temp keys with stripped leading zeros for matching only
            pivot["match_ID"]   = pivot["ID"].astype(str).str.lstrip("0")
            emp_df["match_ID"]  = emp_df["ID"].astype(str).str.lstrip("0")

            pivot = pivot.merge(emp_df[["match_ID", "First Name", "Last Name"]],
                                on="match_ID", how="left")

            pivot.drop(columns="match_ID", inplace=True)
            rename_map = {d: label for d, label in zip(days, fmt_cols)}
            pivot.rename(columns=rename_map, inplace=True)

            # ensure all day columns exist
            for label in fmt_cols:
                  if label not in pivot.columns:
                        pivot[label] = 0.0

            final_cols = ["ID", "First Name", "Last Name"] + fmt_cols
            pivot = pivot[final_cols]

            missing_emp_rows = raw[raw["Number"].isna()]
            st.warning(f"Dropped {len(missing_emp_rows)} rows with missing emp_id.")

            # ── 6) Show table -----------------------------------------------
            gb = GridOptionsBuilder.from_dataframe(pivot)
            gb.configure_default_column(resizable=True, type=["numericColumn"], precision=2)
            for col in ["ID", "First Name", "Last Name"]:
                  gb.configure_column(col, pinned="left", editable=False)

            st.markdown("""
                  <style>
                  .ag-theme-streamlit .ag-root-wrapper {
                        border-radius: 12px !important;
                  }
                  </style>
            """, unsafe_allow_html=True)

            gb.configure_grid_options(domLayout='autoHeight', suppressHorizontalScroll=False)
            gb.configure_grid_options(forceFitColumns=True)

            AgGrid(
                  pivot,
                  gridOptions=gb.build(),
                  theme="streamlit",
                  fit_columns_on_grid_load=True,
                  allow_unsafe_jscode=True,
                  domLayout='autoHeight',
                  key=f"emp_table_{week_start}"
            )
# =============================================== ROOM STATs ================
elif main_choice == "Room STATs":
    import os
    import datetime
    from datetime import date, timedelta
    from dateutil.relativedelta import relativedelta

    st.header("📊 Room STATs")

    # ─────────────────────────── 1. Week navigator ─────────────────────
    if "rs_week_start" not in st.session_state:
        st.session_state.rs_week_start = date.today() + relativedelta(weekday=MO(-1))

    nav_prev, nav_range, nav_next = st.columns([1, 3, 1])

    if nav_prev.button("◀", key="rs_week_prev"):
        st.session_state.rs_week_start -= timedelta(days=7)
    if nav_next.button("▶", key="rs_week_next"):
        st.session_state.rs_week_start += timedelta(days=7)

    week_start = st.session_state.rs_week_start
    week_end   = week_start + timedelta(days=6)
    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    fmt_day    = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
    day_cols   = [d.strftime(fmt_day) for d in week_dates]

    nav_range.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

    # ─────────────────────────── 2. Tab-style Selectors ─────────────────
    subtab_rs = st.tabs(["📈 Actuals", "📅 Forecast", "📊 OTB + Pickup"])
    tab_labels = ["📈 Actuals", "📅 Forecast", "📊 OTB + Pickup"]
    tab_models = [db.RoomActual, db.RoomForecast, db.RoomOTBPickup]

    for tab, label, Model in zip(subtab_rs, tab_labels, tab_models):
        with tab:
            # ─────────────────────── Fetch data ────────────────────────
            rows = (
                session.query(Model)
                       .filter(Model.date.between(week_start, week_end))
                       .all()
            )

            kpis = ["Occupied Rooms", "Arrivals", "Covers", "Departures", "Check-outs"]
            data_dict = {k: {"KPI": k} for k in kpis}

            for r in rows:
                col_lbl = r.date.strftime(fmt_day)
                if r.kpi in data_dict:
                    data_dict[r.kpi][col_lbl] = r.value

            for k in kpis:
                for col in day_cols:
                    data_dict[k].setdefault(col, 0)

            df_edit = pd.DataFrame(data_dict.values()).loc[:, ["KPI"] + day_cols]

            # ───────────── Save Button (Top-right) ─────────────
            save_col = st.columns([10, 1])
            with save_col[1]:
                save_click = st.button("💾", key=f"save_btn_{Model.__name__}")

            # ───────────── Styling ─────────────
            st.markdown("""
                <style>
                section[data-testid="stDataEditor"] thead tr {
                    background-color: #2D2D2D !important;
                }
                section[data-testid="stDataEditor"] thead th {
                    color: white !important;
                    font-weight: bold !important;
                    text-align: center !important;
                    border-right: 1px solid #ddd !important;
                }
                section[data-testid="stDataEditor"] div[data-testid^="cell-"][data-testid$="-0"] {
                    font-weight: bold !important;
                    color: #1F4E79 !important;
                    background-color: #E3F2FD !important;
                }
                section[data-testid="stDataEditor"] div[data-testid^="cell-"] {
                    border-right: 1px solid #eee !important;
                }
                </style>
            """, unsafe_allow_html=True)

            # ───────────── Editable Table ─────────────
            edited_df = st.data_editor(
                df_edit,
                column_config={
                    "KPI": st.column_config.TextColumn("KPI", disabled=True),
                },
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                key=f"room_stats_editor_{Model.__name__}"
            )

            # ───────────── Save Logic ─────────────
            if save_click:
                edited = edited_df.set_index("KPI")

                session.query(Model).filter(
                    Model.date.between(week_start, week_end)
                ).delete(synchronize_session=False)

                for kpi in kpis:
                    for d, col in zip(week_dates, day_cols):
                        session.add(
                            Model(kpi=kpi, date=d, value=int(edited.at[kpi, col]))
                        )

                session.commit()

                st.markdown(f"""
                    <div style="
                        position: fixed;
                        bottom: 24px;
                        right: 24px;
                        background-color: #4CAF50;
                        color: white;
                        padding: 14px 20px;
                        border-radius: 6px;
                        font-size: 15px;
                        box-shadow: 0px 4px 10px rgba(0,0,0,0.25);
                        z-index: 1000;
                        animation: fadeOut 5s forwards;
                    ">
                        ✅ Room stats for <b>{label}</b> saved successfully.
                    </div>
                    <style>
                        @keyframes fadeOut {{
                            0%   {{ opacity: 1; }}
                            80%  {{ opacity: 1; }}
                            100% {{ opacity: 0; display: none; }}
                        }}
                    </style>
                """, unsafe_allow_html=True)
# ---------- SCHEDULING ----------------------------------------------------
elif main_choice == "Scheduling":
    import os, json, pandas as pd
    from datetime import date, timedelta, datetime, time
    from dateutil.relativedelta import relativedelta, MO
    from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
    from sqlalchemy import func
    from collections import defaultdict

    st.header("🗓️ Weekly Scheduling")

    # ---------- 1) week reference ------------------------------------------
    if "sch_week_start" not in st.session_state:
        st.session_state.sch_week_start = date.today() + relativedelta(weekday=MO(-1))

    week_start = st.session_state.sch_week_start
    week_end   = week_start + timedelta(days=6)
    week_dates = [week_start + timedelta(i) for i in range(7)]
    week_dates_str = [d.strftime("%Y-%m-%d") for d in week_dates]
    fmt_day    = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
    day_cols   = [d.strftime(fmt_day) for d in week_dates]

    # ---------- 2) Filters --------------------------------------------------
    emp_df = refresh(db.Employee)
    col_dept, col_pos = st.columns(2)
    with col_dept:
        sel_dept = st.selectbox("Department*", ["(Select)"] + sorted(emp_df["department"].dropna().unique()))
    with col_pos:
        if sel_dept != "(Select)":
            pos_opts = emp_df.loc[emp_df["department"] == sel_dept, "role"].dropna().unique()
            sel_pos  = st.selectbox("Position*", ["(Select)"] + sorted(pos_opts))
            if sel_pos != "(Select)":
                st.session_state["selected_pos"] = sel_pos
            st.session_state["selected_pos"] = sel_pos    # ✅ ADD THIS LINE
            st.session_state["selected_pos"] = sel_pos
        else:
            sel_pos = "(Select)"

    if sel_dept == "(Select)" or sel_pos == "(Select)":
        st.info("Select **Department** and **Position** to view schedule grid.")

        st.stop()

    # ---------- 3) pull employees ------------------------------------------
    emp_sub = emp_df[(emp_df["department"] == sel_dept) & (emp_df["role"] == sel_pos)]
    emp_sub = emp_sub.drop_duplicates(subset=["id"]).reset_index(drop=True)
    if emp_sub.empty:
        st.warning("No employees match that Department / Position.")
        st.stop()

    ids    = emp_sub["name"].str.extract(r"(\d+)$")[0].fillna("")
    firsts = emp_sub["name"].str.extract(r",\s*([^\d]+)")[0].str.strip()
    lasts  = emp_sub["name"].str.extract(r"^\s*([^,]+)")[0].str.strip()

    sched_df = pd.DataFrame({
        "ID": ids,
        "First Name": firsts,
        "Last Name": lasts,
        "emp_id": emp_sub["id"],
    })
    for dc in day_cols:
        sched_df[dc] = ""
    sched_df["Total"] = ""                                  # weekly-hours column
    sched_df = sched_df.drop_duplicates(subset=["ID", "First Name", "Last Name"]).reset_index(drop=True)

    # ---------- 4) load saved rows & OFF highlight -------------------------
    saved_rows = session.query(db.Schedule).filter(
        db.Schedule.emp_id.in_(emp_sub["id"]),
        db.Schedule.day.in_(week_dates_str)
    ).all()

    avail_rows = session.query(db.Schedule).filter(
        db.Schedule.emp_id.in_(emp_sub["id"]),
        db.Schedule.day.in_(WK_TO_DATE.values())
    ).all()

    off_map = {}  # emp_id → set of weekday abbreviations like 'Mon'
    for r in avail_rows:
        if r.shift_type.upper() == "OFF":
            lbl = r.day.strftime("%a")
            off_map.setdefault(r.emp_id, set()).add(lbl)

    for r in saved_rows:
        col = r.day.strftime(fmt_day)
        idx = sched_df.index[sched_df["emp_id"] == r.emp_id][0]
        sched_df.at[idx, col] = r.shift_type

    cell_styles = {}
    for idx, row in sched_df.iterrows():
        for d, col in zip(week_dates, day_cols):
            if row["emp_id"] in off_map and d.strftime("%a") in off_map[row["emp_id"]]:
                sched_df.at[idx, col] = "OFF"
                cell_styles[f"{row['ID']}|||{col}"] = {"backgroundColor": "#f8d7da"}
    # ---------- helpers -----------------------------------------------------
    def do_rerun():
        st.rerun() if hasattr(st, "rerun") else st.experimental_rerun()

    # keep one-step undo snapshot in session_state so it survives rerun
    if "undo_backup" not in st.session_state:
        st.session_state["undo_backup"] = {}

    def backup_diffs(old_map, _new_map):
        """Save old_map (state before last save) for a one-step undo."""
        import copy
        st.session_state["undo_backup"] = copy.deepcopy(old_map)

    def undo_last_change():
        undo_map = st.session_state.get("undo_backup", {})
        if not undo_map:
            st.warning("Nothing to undo.")
            return

        # Ensure day values are proper date objects
        emp_ids  = {k[0] for k in undo_map}
        day_vals = {k[1] if isinstance(k[1], date) else datetime.strptime(k[1], "%Y-%m-%d").date()
                    for k in undo_map}

        session.query(db.Schedule).filter(
            db.Schedule.emp_id.in_(emp_ids),
            db.Schedule.day.in_(day_vals)
        ).delete(synchronize_session=False)

        session.bulk_save_objects([
            db.Schedule(
                emp_id=k[0],
                day=k[1] if isinstance(k[1], date) else datetime.strptime(k[1], "%Y-%m-%d").date(),
                shift_type=v
            )
            for k, v in undo_map.items()
        ])
        session.commit()

        st.session_state["undo_backup"] = {}
        st.success("Undo complete.")
        do_rerun()
    # ---------- 5) shift list + quick add -----------------------------------
    st.markdown("#### Shifts available for this position")
    ShiftTbl = db.ShiftTime
    pos_rec  = session.query(db.Position).filter(func.lower(db.Position.name)==sel_pos.lower()).first()
    shift_rows = session.query(ShiftTbl).filter(ShiftTbl.position_id==pos_rec.id).all()
    fmt = lambda t: t.strftime("%H:%M") if t else ""
    shift_strings = sorted({f"{fmt(r.start)}-{fmt(r.end)}" for r in shift_rows if r.start and r.end})
    buckets = {"Morning": [], "Afternoon": [], "Evening": []}
    for r in shift_rows:
        buckets[r.period].append(f"{fmt(r.start)}-{fmt(r.end)}")
    for p in ["Morning", "Afternoon", "Evening"]:
        st.markdown(f"**{p}:** {', '.join(sorted(buckets[p])) if buckets[p] else '*No shifts created*'}")
    shift_opts = shift_strings[:]

    with st.expander("➕ Add new shift", expanded=False):
        period = st.selectbox("Time of day", ["Morning", "Afternoon", "Evening"])
        c1, c2 = st.columns(2)
        with c1: t_start = st.time_input("Start", value=time(0,0), step=1800)
        with c2: t_end   = st.time_input("End",   value=time(12,0), step=1800)
        if st.button("💾 Save shift"):
            if t_end <= t_start:
                st.error("End time must be after start.")
            else:
                new_shift = f"{t_start.strftime('%H:%M')}-{t_end.strftime('%H:%M')}"
                if new_shift in shift_opts:
                    st.warning("That shift already exists.")
                else:
                    session.add(ShiftTbl(position_id=pos_rec.id, period=period, start=t_start, end=t_end))
                    session.commit(); st.success("Shift saved."); do_rerun()

    # ---------- 6) editable grid -------------------------------------------
    df_view = sched_df.copy()

    # --- dynamic Total hours -----------------------------------------------
    total_js = JsCode(f"""
        function(p) {{
            const cols = {json.dumps(day_cols)};
            let t = 0;
            const hh = s => {{
                if(!s || s.toUpperCase()==="OFF") return 0;
                const a = s.split("-");
                if(a.length !== 2) return 0;
                const [h0,m0] = a[0].split(":").map(Number);
                const [h1,m1] = a[1].split(":").map(Number);
                let diff = (h1*60+m1) - (h0*60+m0);
                if (diff <= 0) diff += 1440;
                return diff/60;
            }};
            cols.forEach(c => t += hh(p.data[c] || ""));
            return t.toFixed(2);
        }}
    """)

    # --- renderer (no raw HTML) & class rules -------------------------------
    icon_renderer = JsCode("""
        function(p) {
            const v = Number(p.value || 0).toFixed(2);
            const icon = v > 40 ? " ⚠️" : "";
            return v + icon;
        }
    """)

    # --- bold / red styling for Total ---------------------------------------
    style_total = JsCode("""
        function(p) {
            const v = Number(p.value || 0);
            if (v > 40) {
                return {fontWeight: 'bold', color: '#d9534f'};
            }
            return {fontWeight: 'bold'};
        }
    """)

    gb = GridOptionsBuilder.from_dataframe(df_view)
    gb.configure_default_column(editable=True)

    for col in ["ID", "First Name", "Last Name", "emp_id"]:
        gb.configure_column(col, editable=False, hide=True if col=="emp_id" else False)

    gb.configure_grid_options(enableRangeSelection=True, enableFillHandle=True,
                              undoRedoCellEditing=True, undoRedoCellEditingLimit=100,
                              clipboardDelimiter=",", domLayout="normal")
    gb.configure_grid_options(getContextMenuItems=JsCode("""
        function(p){return ['copy','copyWithHeaders','paste','undo','redo'];}
    """))

    js_style = JsCode(f"""
        function(p){{
            const m = {json.dumps(cell_styles)};
            const k = p.data['ID'] + '|||' + p.colDef.field;
            return m[k] || null;
        }}
    """)

    for col in day_cols:
        gb.configure_column(col,
                            cellEditor="agSelectCellEditor",
                            cellEditorParams={"values": shift_opts + ["OFF", ""]},
                            cellStyle=js_style,
                            flex=1)

    # -- Total column (bold, warn >40) --------------------------------------
    gb.configure_column(
        "Total",
        editable=False,
        valueGetter=total_js,
        cellRenderer=icon_renderer,
        cellStyle=style_total,     # ← applies bold (and red if >40 h)
        flex=1
    )

    # ---------- navigator & buttons ----------------------------------------
    nav_prev2, nav_range2, nav_next2 = st.columns([1,3,1])
    if nav_prev2.button("◀", key="sch_prev2"):
        st.session_state.sch_week_start -= timedelta(days=7); do_rerun()
    if nav_next2.button("▶", key="sch_next2"):
        st.session_state.sch_week_start += timedelta(days=7); do_rerun()
    nav_range2.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

    hdr_l, hdr_u, hdr_s, hdr_c = st.columns([6,1,1,2])
    with hdr_u: undo_clicked  = st.button("🔄 Undo")
    with hdr_s: save_clicked  = st.button("💾 Save")
    with hdr_c: copy_clicked  = st.button("📋 Copy Forward")

    # ---------- grid CSS ----------------------------------------------------
    st.markdown("""
        <style>
        .ag-theme-alpine .ag-cell,
        .ag-theme-alpine .ag-header-cell-label{
            font-size:13px!important;padding:4px 8px!important;
        }
        .ag-theme-alpine .ag-header-cell-label{
            font-weight:bold;justify-content:center!important;
        }
        .ag-theme-alpine .ag-cell,
        .ag-theme-alpine .ag-header-cell{
            box-shadow:inset -1px 0 #b3b3b3,inset 1px 0 #b3b3b3;
        }
        .ag-theme-alpine .ag-row,
        .ag-theme-alpine .ag-header-row{
            border-bottom:1px solid #dcdcdc!important;
        }
        .ag-theme-alpine .ag-root-wrapper{
            border:1px solid #a0a0a0;border-radius:4px;
        }
        /* NEW — styling for Total column */
        .ag-theme-alpine .total-normal{font-weight:bold;}
        .ag-theme-alpine .total-over  {font-weight:bold;color:#d9534f;}
        </style>
    """, unsafe_allow_html=True)

    st.markdown("""
        <style>
        .ag-theme-streamlit .ag-cell,
        .ag-theme-streamlit .ag-header-cell-label {
            font-size:13px!important; padding:4px 8px!important;
        }
        .ag-theme-streamlit .ag-header-cell-label {
            font-weight:bold; justify-content:center!important;
        }
        .ag-theme-streamlit .ag-cell,
        .ag-theme-streamlit .ag-header-cell {
            box-shadow:inset -1px 0 #b3b3b3,inset 1px 0 #b3b3b3;
        }
        .ag-theme-streamlit .ag-row,
        .ag-theme-streamlit .ag-header-row {
            border-bottom:1px solid #dcdcdc!important;
        }
        .ag-theme-streamlit .ag-root-wrapper {
            border:1px solid #a0a0a0; border-radius:6px;
        }
        </style>
    """, unsafe_allow_html=True)

    grid_response = AgGrid(
        df_view,
        gridOptions=gb.build(),
        theme="streamlit",
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=True,
        domLayout="autoHeight",
        data_return_mode="AS_INPUT",
        update_mode="MODEL_CHANGED"
    )
    edited_df = grid_response["data"]
    # Ensure selected_pos and pos_id are available
    try:
        selected_pos = selected_pos
    except NameError:
        selected_pos = None

    pos_id = (
        session.query(Position.id)
        .filter_by(name=selected_pos)
        .scalar()
        if selected_pos else None
    )

    # ---------- PLANNING PERIOD SUMMARY ------------------------------------
    def parse_shift_hours(s):
        try:
            if not s or s.upper() == "OFF": return 0.0
            a,b = s.split("-")
            t0 = datetime.strptime(a, "%H:%M")
            t1 = datetime.strptime(b, "%H:%M")
            h  = (t1 - t0).seconds / 3600
            return round(h + 24 if h <= 0 else h, 2)
        except:
            return 0.0

    summary = defaultdict(lambda: {"Shifts":0, "Total":0.0})
    for col in day_cols:
        for _, r in edited_df.iterrows():
            s = str(r[col]).strip()
            if s and s.upper() != "OFF":
                summary[col]["Shifts"] += 1
                summary[col]["Total"]  += parse_shift_hours(s)

    st.markdown("#### 📊 Planning Period Summary")
    summary_df = pd.DataFrame({
        "Metric": ["Shifts", "Total Hours"],
        **{c: [summary[c]["Shifts"], f'{summary[c]["Total"]:.2f}'] for c in day_cols}
    }).set_index("Metric")
    st.dataframe(summary_df, height=120, use_container_width=True)

    visible_positions = [selected_pos] if selected_pos and selected_pos != "(Select)" else []
    st.session_state["planning_summary_all"] = st.session_state.get("planning_summary_all", {})
    for pos in visible_positions:
        st.session_state["planning_summary_all"][pos] = summary_df

    selected_pos = st.session_state.get("selected_pos")

    # ---------- PROJECTED / STANDARD HOURS (Totals Only) ----------------------
    if not selected_pos or selected_pos == "(Select)":
        st.info("Select a position above to see projected hours.")
        st.stop()

    std_rows = (
        session.query(LaborStandard)
        .join(Position, LaborStandard.position_id == Position.id)
        .filter(Position.name == selected_pos)
        .all()
    )
    if not std_rows:
        st.info("No labor standards defined for this position.")
        st.stop()

    hrs_per_fte    = 8
    week_start     = st.session_state.get("sch_week_start", date.today())
    week_dates     = [week_start + timedelta(days=i) for i in range(7)]
    fmt_day        = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
    col_labels     = [d.strftime(fmt_day) for d in week_dates]

    grand_tot = {lbl: {"Shifts": 0.0, "Hours": 0.0} for lbl in col_labels}
    for std in std_rows:
        metric      = std.metric
        std_per_fte = std.standard
        fc_rows = (
            session.query(RoomForecast)
            .filter(RoomForecast.kpi.ilike(metric))
            .filter(RoomForecast.date.in_(week_dates))
            .all()
        )
        fc_map = {(r.date): r.value for r in fc_rows}
        for d, lbl in zip(week_dates, col_labels):
            rooms   = fc_map.get(d, 0)
            shifts  = round(rooms / std_per_fte, 2) if rooms else 0.0
            hours   = round(shifts * hrs_per_fte, 2)
            grand_tot[lbl]["Shifts"] += shifts
            grand_tot[lbl]["Hours"]  += hours

    df_tot = pd.DataFrame({
        "Metric": ["Total Shifts", "Total Hours"],
        **{
            lbl: [
                f'{grand_tot[lbl]["Shifts"]:.2f}',
                f'{grand_tot[lbl]["Hours"]:.2f}'
            ] for lbl in col_labels
        }
    }).set_index("Metric")

    st.markdown("#### 📊 Projected / Standard Hours")
    st.dataframe(df_tot, height=120, use_container_width=True)

    st.session_state["projected_hours_all"] = st.session_state.get("projected_hours_all", {})
    for pos in visible_positions:
        st.session_state["projected_hours_all"][pos] = df_tot

    # ---------- OTB + Pickup Projected Hours (Totals Only) ----------------------
    otb_tot = {lbl: {"Shifts": 0.0, "Hours": 0.0} for lbl in col_labels}
    for std in std_rows:
        metric      = std.metric
        std_per_fte = std.standard
        otb_rows = (
            session.query(db.RoomOTBPickup)
            .filter(db.RoomOTBPickup.kpi.ilike(metric))
            .filter(db.RoomOTBPickup.date.in_(week_dates))
            .all()
        )
        otb_map = {r.date: r.value for r in otb_rows}
        for d, lbl in zip(week_dates, col_labels):
            rooms   = otb_map.get(d, 0)
            shifts  = round(rooms / std_per_fte, 2) if rooms else 0.0
            hours   = round(shifts * hrs_per_fte, 2)
            otb_tot[lbl]["Shifts"] += shifts
            otb_tot[lbl]["Hours"]  += hours

    df_otb = pd.DataFrame({
        "Metric": ["Total Shifts", "Total Hours"],
        **{
            lbl: [
                f'{otb_tot[lbl]["Shifts"]:.2f}',
                f'{otb_tot[lbl]["Hours"]:.2f}'
            ] for lbl in col_labels
        }
    }).set_index("Metric")

    st.markdown("#### 📊 OTB + Pickup Projected Hours")
    st.dataframe(df_otb, height=120, use_container_width=True)

    st.session_state["otb_hours_all"] = st.session_state.get("otb_hours_all", {})
    for pos in visible_positions:
        st.session_state["otb_hours_all"][pos] = df_otb   
 # ---------- SAVE --------------------------------------------------------
    if save_clicked:
        old_rows = session.query(db.Schedule).filter(
            db.Schedule.emp_id.in_(sched_df["emp_id"]),
            db.Schedule.day.in_(week_dates_str)
        ).all()
        old_map = {(r.emp_id, r.day): r.shift_type for r in old_rows}

        new_map = {}
        for _, r in edited_df.iterrows():
            emp = int(r["emp_id"])
            for d, col in zip(week_dates, day_cols):
                val = str(r[col]).strip()
                if val:
                    new_map[(emp, d.strftime("%Y-%m-%d"))] = val

        backup_diffs(old_map, new_map)

        # remove old rows for this week
        session.query(db.Schedule).filter(
            db.Schedule.emp_id.in_(sched_df["emp_id"]),
            db.Schedule.day.in_(week_dates_str)
        ).delete(synchronize_session=False)

        # ---------- FIX: convert str → date before insert ------------------
        from datetime import datetime
        new_rows = [
            db.Schedule(
                emp_id=k[0],
                day=datetime.strptime(k[1], "%Y-%m-%d").date(),  # ← convert here
                shift_type=v
            )
            for k, v in new_map.items()
        ]
        if new_rows:
            session.bulk_save_objects(new_rows)
            session.commit()

        st.success("Schedule saved.")
    # ---------- UNDO --------------------------------------------------------
    if undo_clicked:
        undo_last_change()

    # ---------- COPY FORWARD (always overwrite target week) -----------------
    if copy_clicked:
        src_rows = session.query(db.Schedule).filter(
            db.Schedule.emp_id.in_(sched_df["emp_id"]),
            db.Schedule.day.in_(week_dates_str)
        ).all()

        if not src_rows:
            st.info("There is nothing saved for the current week yet.")
        else:
            tgt_dates = [d + timedelta(days=7) for d in week_dates]

            session.query(db.Schedule).filter(
                db.Schedule.emp_id.in_(sched_df["emp_id"]),
                db.Schedule.day.in_(tgt_dates)
            ).delete(synchronize_session=False)

            new_rows = [
                db.Schedule(
                    emp_id=r.emp_id,
                    day=(r.day if isinstance(r.day, date) else datetime.strptime(r.day, "%Y-%m-%d").date()) + timedelta(days=7),
                    shift_type=r.shift_type
                )
                for r in src_rows
            ]
            session.bulk_save_objects(new_rows)
            session.commit()
            st.success("Copied current week to next week and overwrote any existing data.")

# ------------------ PAGE: COST AND OT MGMT ------------------
elif main_choice == "Cost and OT Mgmt":
      tab1, tab2 = st.tabs(["🚨 OT Risk", "📊 Cost Mgmt"])

      with tab1:
            st.header("⚠️ Overtime Risk Overview")

            # ─────────────── Filters + Week navigator ───────────────
            dept_df = refresh(db.Department).rename(columns={"name": "dept"})
            pos_df  = (
                  refresh(db.Position)
                  .merge(dept_df, left_on="department_id", right_on="id")
                  [["id_x", "name", "dept"]]
                  .rename(columns={"id_x": "id"})
            )

            dept_opts = ["(All)"] + sorted(dept_df["dept"].dropna().unique())
            col1, col2 = st.columns(2)
            with col1:
                  sel_dept = st.selectbox("Department", dept_opts, key="otrisk_dept")
            if sel_dept != "(All)":
                  pos_opts = ["(All)"] + sorted(
                        pos_df.loc[pos_df["dept"] == sel_dept, "name"].dropna().unique()
                  )
            else:
                  pos_opts = ["(All)"]
            with col2:
                  sel_pos = st.selectbox("Position", pos_opts, key="otrisk_pos")

            if "otrisk_week_start" not in st.session_state:
                  st.session_state.otrisk_week_start = date.today() + relativedelta(weekday=MO(-1))

            col_prev, col_range, col_next = st.columns([1, 3, 1])
            if col_prev.button("◀", key="otrisk_prev"):
                  st.session_state.otrisk_week_start -= timedelta(days=7)
            if col_next.button("▶", key="otrisk_next"):
                  st.session_state.otrisk_week_start += timedelta(days=7)

            week_start = st.session_state.otrisk_week_start
            week_end   = week_start + timedelta(days=6)
            col_range.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

            # ─────────────── Query actual hours for the week ───────────────
            q = (
                  session.query(
                        db.Actual.emp_id.label("Number"),
                        db.Actual.date.label("Business Date"),
                        (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
                        db.Position.name.label("Position"),
                        db.Department.name.label("Department")
                  )
                  .join(db.Position,   db.Actual.position_id == db.Position.id)
                  .join(db.Department, db.Position.department_id == db.Department.id)
                  .filter(db.Actual.date.between(week_start, week_end))
                  .filter(or_(db.Actual.hours != 0, db.Actual.ot_hours != 0))
            )

            if sel_dept != "(All)":
                  q = q.filter(db.Department.name == sel_dept)
            if sel_pos != "(All)":
                  q = q.filter(db.Position.name == sel_pos)

            raw = pd.DataFrame(q.all(),
                               columns=["Number", "Business Date", "Hours", "Position", "Department"])

            if raw.empty:
                  st.warning("No actual hours data found for selected filters.")
                  st.stop()

            raw["Number"] = (
                  pd.to_numeric(raw["Number"], errors="coerce")
                    .fillna(0)
                    .astype(int)
                    .astype(str)
                    .str.zfill(5)
            )

            # Merge with employee names
            emp_df = refresh(db.Employee).copy()
            parts = emp_df["name"].astype(str).str.extract(
                  r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
            )
            emp_df["ID"]         = parts["ID"].fillna("").astype(str).str.strip().str.zfill(5)
            emp_df["First Name"] = parts["First_Name"].str.strip()
            emp_df["Last Name"]  = parts["Last_Name"].str.strip()
            emp_df["match_ID"]   = emp_df["ID"].astype(str).str.lstrip("0")
            raw["match_ID"]      = raw["Number"].astype(str).str.lstrip("0")
            merged = raw.merge(emp_df[["match_ID", "First Name", "Last Name"]],
                               on="match_ID", how="left")

            # ─────────────── Aggregate actuals ───────────────
            agg = merged.groupby(["Number", "First Name", "Last Name"]).agg(
                  total_hours=("Hours", "sum"),
                  days_worked=("Business Date", pd.Series.nunique)
            ).reset_index()

            # ─────────────── Pull Scheduled Days ───────────────
            sched_rows = (
                  session.query(db.Employee.name, db.Schedule.day, db.Schedule.shift_type)
                  .join(db.Employee, db.Employee.id == db.Schedule.emp_id)
                  .filter(db.Schedule.day.between(week_start, week_end))
                  .all()
            )
            sched_df = pd.DataFrame(sched_rows, columns=["name", "day", "shift_type"])

            if not sched_df.empty:
                  sched_df["shift_type"] = sched_df["shift_type"].fillna("").astype(str).str.upper().str.strip()
                  sched_df = sched_df[sched_df["shift_type"] != "OFF"]

                  sched_df["Number"] = sched_df["name"].str.extract(r"(\d+)$")[0].fillna("").str.zfill(5)
                  sched_df["day"] = pd.to_datetime(sched_df["day"])
                  merged["Business Date"] = pd.to_datetime(merged["Business Date"])

                  last_worked = (
                        merged.groupby("Number")["Business Date"]
                        .max()
                        .reset_index()
                        .rename(columns={"Business Date": "last_worked"})
                  )

                  sched_df = sched_df.merge(last_worked, on="Number", how="left")
                  sched_df["after_work"] = sched_df["day"] > sched_df["last_worked"]

                  sched_counts = sched_df.groupby("Number")["day"].nunique().reset_index(name="Days Scheduled")

                  sched_future = sched_df[sched_df["after_work"]].copy()
                  days_remaining = sched_future.groupby("Number")["day"].nunique().reset_index(name="Days Remaining")

                  def parse_shift_to_hours(shift_str):
                        try:
                              start, end = shift_str.split("-")
                              start_dt = pd.to_datetime(start, format="%H:%M")
                              end_dt = pd.to_datetime(end, format="%H:%M")
                              hours = (end_dt - start_dt).total_seconds() / 3600
                              if hours < 0:
                                    hours += 24
                              return max(0, hours - 0.5)
                        except:
                              return 0

                  sched_future["shift_hours"] = sched_future["shift_type"].apply(parse_shift_to_hours)
                  future_hours = sched_future.groupby("Number")["shift_hours"].sum().reset_index()
                  future_hours.rename(columns={"shift_hours": "Future Scheduled Hrs"}, inplace=True)

                  agg = agg.merge(sched_counts, how="left", on="Number")
                  agg = agg.merge(days_remaining, how="left", on="Number")
                  agg = agg.merge(future_hours, how="left", on="Number")

            else:
                  agg["Days Scheduled"] = 0
                  agg["Days Remaining"] = 0
                  agg["Future Scheduled Hrs"] = 0

            agg["Days Scheduled"] = agg["Days Scheduled"].fillna(0).astype(int)
            agg["Days Remaining"] = agg["Days Remaining"].fillna(0).astype(int)
            agg["Future Scheduled Hrs"] = agg["Future Scheduled Hrs"].fillna(0)

            agg.rename(columns={"days_worked": "Days Worked"}, inplace=True)
            agg["OT Risk"] = agg["total_hours"].apply(lambda h: "No Risk" if h <= 40 else "At Risk")
            agg["OT Risk %"] = agg["total_hours"].apply(
                  lambda h: "0%" if pd.isna(h) or h <= 40 else f"{round(((h - 40)/40)*100)}%"
            )
            agg["Projected OT"] = agg["total_hours"].apply(lambda h: max(round(h - 40, 2), 0))
            agg["Future Scheduled Hrs"] = pd.to_numeric(agg["Future Scheduled Hrs"], errors="coerce").fillna(0)
            agg["Total Hrs Worked + Schedule"] = (agg["total_hours"] + agg["Future Scheduled Hrs"]).round(2)

            def classify_ot_risk(row):
                  if row["Total Hrs Worked + Schedule"] <= 40:
                        return "No Risk"
                  elif row["Days Remaining"] > 0:
                        return "At Risk"
                  else:
                        return "OT"

            def estimate_risk_percent(row):
                  if row["Total Hrs Worked + Schedule"] <= 40:
                        return "0%"
                  if row["Days Remaining"] == 0:
                        return "100%"
                  elif row["Days Remaining"] == 1:
                        return "80%"
                  elif row["Days Remaining"] == 2:
                        return "60%"
                  elif row["Days Remaining"] == 3:
                        return "40%"
                  else:
                        return "20%"

            agg["OT Risk"] = agg.apply(classify_ot_risk, axis=1)
            agg["OT Risk %"] = agg.apply(estimate_risk_percent, axis=1)
            agg["Projected OT"] = agg["Total Hrs Worked + Schedule"].apply(lambda h: max(round(h - 40, 2), 0))

            # ─────────────── Merge Employee Rate + OT Cost ───────────────
            emp_df = refresh(db.Employee).copy()
            emp_df["ID"] = (
                  emp_df["name"]
                  .astype(str)
                  .str.extract(r"(\d+)$")[0]
                  .fillna("")
                  .str.zfill(5)
            )

            # Ensure consistent string types before merge
            emp_df["ID"] = emp_df["ID"].astype(str)

            if "hourly_rate" in emp_df.columns:
                  emp_df["rate"] = emp_df["hourly_rate"].fillna(0)
            else:
                  emp_df["rate"] = 0.00

            # Coerce Number to string too
            agg["Number"] = agg["Number"].astype(str)

            # Merge safely
            agg = agg.merge(emp_df[["ID", "rate"]], left_on="Number", right_on="ID", how="left")
            agg["rate"] = agg["rate"].fillna(0)
            agg["OT Cost"] = (agg["Projected OT"] * agg["rate"] * 1.5).round(2)
            # ─────────────── Display Table ───────────────
            show_cols = [
                  "Number", "First Name", "Last Name",
                  "OT Risk", "OT Risk %", "Projected OT", "OT Cost",
                  "Days Worked", "Days Remaining",
                  "Total Hrs Worked + Schedule"
            ]

            from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

            risk_icon_renderer = JsCode("""
                  function(params) {
                        if (params.value === "OT") {
                              return "⛔ OT";
                        } else if (params.value === "At Risk") {
                              return "⚠️ At Risk";
                        } else {
                              return "✅ No Risk";
                        }
                  }
            """)

            risk_pct_renderer = JsCode("""
                  function(params) {
                        let pct = parseFloat(params.value.replace('%', ''));
                        if (pct >= 100) {
                              return "🔴 " + params.value;  // High risk
                        } else if (pct >= 80) {
                              return "🟤 " + params.value;  // Brown for very high approaching risk
                        } else if (pct >= 40) {
                              return "🟠 " + params.value;  // Medium risk
                        } else if (pct > 0) {
                              return "🟡 " + params.value;  // Light risk
                        } else {
                              return "🟢 " + params.value;  // No risk
                        }
                  }
            """)

            currency_renderer = JsCode("""
                  function(params) {
                        if (!params.value || isNaN(params.value)) return "";
                        return "$" + Number(params.value).toFixed(2);
                  }
            """)

            gb = GridOptionsBuilder.from_dataframe(agg[show_cols])
            gb.configure_column("OT Risk", cellRenderer=risk_icon_renderer)
            gb.configure_column("OT Risk %", cellRenderer=risk_pct_renderer)
            gb.configure_column("OT Cost", cellRenderer=currency_renderer, type=["numericColumn"])
            gb.configure_default_column(editable=False, filter=True, resizable=True)

            AgGrid(
                  agg[show_cols],
                  gridOptions=gb.build(),
                  fit_columns_on_grid_load=True,
                  height=420,
                  allow_unsafe_jscode=True,
                  enable_enterprise_modules=False
            )

            st.session_state["ot_risk_final"] = agg.copy()
            st.session_state["ot_risk_filters"] = {
                  "week_start": week_start,
                  "week_end": week_end,
                  "department": sel_dept,
                  "position": sel_pos
            }

# ─────────────── Cost Mgmt Table ───────────────
      with tab2:
            st.header("💸 Cost Management: FTE Variance")

            from collections import defaultdict

            dept_df = refresh(db.Department).rename(columns={"name": "dept"})
            dept_list = sorted(dept_df["dept"].dropna().unique())
            sel_dept = st.selectbox("Select Department", dept_list)

            if "cost_week_start" not in st.session_state:
                  st.session_state.cost_week_start = date.today() + relativedelta(weekday=MO(-1))

            prev_col, mid_col, next_col = st.columns([1,3,1])
            if prev_col.button("⬅", key="cost_prev"):
                  st.session_state.cost_week_start -= timedelta(days=7)
            if next_col.button("➡", key="cost_next"):
                  st.session_state.cost_week_start += timedelta(days=7)

            week_start = st.session_state.cost_week_start
            week_end   = week_start + timedelta(days=6)
            week_dates = [week_start + timedelta(days=i) for i in range(7)]
            fmt_day    = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
            day_cols   = [d.strftime(fmt_day) for d in week_dates]
            mid_col.markdown(f"### Week of {week_start:%b %d, %Y}")

            # ─────────────── Actual Hours ───────────────
            actual_q = (
                  session.query(
                        db.Position.name.label("Position"),
                        db.Actual.date.label("Date"),
                        func.sum(db.Actual.hours).label("Actual Hours")
                  )
                  .join(db.Position, db.Actual.position_id == db.Position.id)
                  .join(db.Department, db.Position.department_id == db.Department.id)
                  .filter(db.Department.name == sel_dept)
                  .filter(db.Actual.date.between(week_start, week_end))
                  .group_by(db.Position.name, db.Actual.date)
            )
            actual_df = pd.DataFrame(actual_q.all())
            if not actual_df.empty:
                  actual_df["FTE"] = actual_df["Actual Hours"] / 8
                  actual_df["Date Label"] = pd.to_datetime(actual_df["Date"]).dt.strftime(fmt_day)
            else:
                  actual_df = pd.DataFrame(columns=["Position", "Date", "Actual Hours", "FTE", "Date Label"])

            # ─────────────── Scheduled FTEs ───────────────
            sched_q = (
                  session.query(
                        db.Employee.role.label("Position"),
                        db.Schedule.day,
                        db.Schedule.shift_type
                  )
                  .join(db.Employee, db.Employee.id == db.Schedule.emp_id)
                  .filter(db.Employee.department == sel_dept)
                  .filter(db.Schedule.day.between(week_start, week_end))
            )
            sched_df_raw = pd.DataFrame(sched_q.all())

            def parse_hours(shift):
                  if not shift or shift.upper() == "OFF": return 0
                  try:
                        a,b = shift.split("-")
                        t0 = datetime.strptime(a.strip(), "%H:%M")
                        t1 = datetime.strptime(b.strip(), "%H:%M")
                        diff = (t1 - t0).seconds / 3600
                        return diff if diff > 0 else diff + 24
                  except:
                        return 0

            if not sched_df_raw.empty:
                  sched_df_raw["Hours"] = sched_df_raw["shift_type"].apply(parse_hours)
                  sched_df_raw["Date Label"] = pd.to_datetime(sched_df_raw["day"]).dt.strftime(fmt_day)
                  sched_group = sched_df_raw.groupby(["Position", "Date Label"])["Hours"].sum().reset_index()
                  sched_group["FTE"] = sched_group["Hours"] / 8
            else:
                  sched_group = pd.DataFrame(columns=["Position", "Date Label", "Hours", "FTE"])

            # ─────────────── OTB FTEs ───────────────
            std_q = (
                  session.query(
                        db.Position.name.label("Position"),
                        db.LaborStandard.metric,
                        db.LaborStandard.standard
                  )
                  .join(db.Position, db.Position.id == db.LaborStandard.position_id)
                  .join(db.Department, db.Position.department_id == db.Department.id)
                  .filter(db.Department.name == sel_dept)
            )
            std_df = pd.DataFrame(std_q.all())

            otb_q = (
                  session.query(
                        db.RoomOTBPickup.date.label("date"),
                        db.RoomOTBPickup.kpi.label("kpi"),
                        db.RoomOTBPickup.value.label("value")
                  )
                  .filter(db.RoomOTBPickup.date.between(week_start, week_end))
            )
            otb_df_raw = pd.DataFrame(otb_q.all())
            if otb_df_raw.empty:
                  otb_df_raw = pd.DataFrame(columns=["date", "kpi", "value"])

            otb_rows = []
            for _, row in std_df.iterrows():
                  pos, metric, std_val = row["Position"], row["metric"], row["standard"]
                  subset = otb_df_raw[otb_df_raw["kpi"].str.lower() == metric.lower()]
                  for _, r in subset.iterrows():
                        date_lbl = pd.to_datetime(r["date"]).strftime(fmt_day)
                        fte = (r["value"] / std_val) if std_val else 0
                        otb_rows.append({
                              "Position": pos,
                              "Date Label": date_lbl,
                              "OTB FTE": fte
                        })
            otb_df = pd.DataFrame(otb_rows)
            if otb_df.empty:
                  otb_df = pd.DataFrame(columns=["Position", "Date Label", "OTB FTE"])

            # ─────────────── No Data Message ───────────────
            if actual_df.empty and sched_group.empty and otb_df.empty:
                  st.warning("No Actual, Scheduled, or OTB data available for the selected week and department.")

            # ─────────────── Final Variance Table ───────────────
            pos_list = sorted(set(actual_df.get("Position", pd.Series()))
                              .union(sched_group.get("Position", pd.Series()))
                              .union(otb_df.get("Position", pd.Series())))

            data = []
            for pos in pos_list:
                  row = {"Position": pos}
                  for d in week_dates:
                        lbl = d.strftime(fmt_day)
                        otb_fte = otb_df[(otb_df["Position"] == pos) & (otb_df["Date Label"] == lbl)]["OTB FTE"].sum()
                        actual_fte = actual_df[(actual_df["Position"] == pos) & (actual_df["Date Label"] == lbl)]["FTE"].sum()
                        sched_fte  = sched_group[(sched_group["Position"] == pos) & (sched_group["Date Label"] == lbl)]["FTE"].sum()

                        if actual_fte > 0:
                              fte_var = actual_fte - otb_fte
                              color = "#ffe6cc"
                              font_color = "red"
                              arrow = "🔺"
                        else:
                              fte_var = sched_fte - otb_fte
                              color = "#e6f0ff"
                              font_color = "black"
                              arrow = "🔽" if fte_var < 0 else ""

                        cell = f"<div style='color:{font_color};background:{color};padding:4px;border-radius:4px;text-align:center'>{arrow} {fte_var:.2f}</div>"
                        row[lbl] = cell
                  data.append(row)

            var_df = pd.DataFrame(data)

            st.markdown("""
            <style>
                  .fancy-table td {
                        text-align: center;
                        vertical-align: middle;
                  }
            </style>
            """, unsafe_allow_html=True)

            st.write(var_df.to_html(escape=False, index=False, classes="fancy-table"), unsafe_allow_html=True)

            # ─────────────── Debug Outputs ───────────────
            with st.expander("🔍 Actual FTE DataFrame", expanded=False):
                  st.dataframe(actual_df, use_container_width=True)

            with st.expander("🔍 Scheduled FTE DataFrame", expanded=False):
                  st.dataframe(sched_group, use_container_width=True)

            with st.expander("🔍 OTB FTE DataFrame", expanded=False):
                  st.dataframe(otb_df, use_container_width=True)


# ─────────────── Reports Page ───────────────
elif main_choice == "Reports":
      from datetime import date, timedelta
      from dateutil.relativedelta import relativedelta, MO
      import pandas as pd
      import io
      from sqlalchemy import or_

      st.header("📋 Reports")

      # ─────────────── Report Type Selection ───────────────
      report_type = st.selectbox(
            "🧾 Select Report Type",
            [
                  "Department Schedule",
                  "Labor Variance",
                  "Forecast Variance",
                  "OT Risk",
                  "Productivity Index",
                  "Labor Standards",
                  "Schedule Variance",
            ]
      )

      # ─────────────── Export Options ───────────────
      st.markdown("### 📤 Export Options")
      export_format = st.radio("Choose Export Format", ["Excel", "PDF", "CSV"], horizontal=True)

      # ─────────────── OT RISK REPORT ───────────────
      if report_type == "OT Risk":
            st.markdown("### 🔍 OT Risk Filters")

            if "ot_risk_ref" not in st.session_state:
                  st.session_state.ot_risk_ref = date.today()

            selected_date = st.date_input(
                  "Select any date in the week",
                  value=st.session_state.ot_risk_ref,
                  key="ot_risk_date_input"
            )

            week_start = selected_date + relativedelta(weekday=MO(-1))
            week_end = week_start + timedelta(days=6)
            st.markdown(f"📅 **Selected Week (OT Risk):** {week_start} to {week_end}")

            dept_df = refresh(db.Department).rename(columns={"name": "Department"})
            dept_list = sorted(dept_df["Department"].dropna().unique())
            ot_risk_dept = st.selectbox("Select Department (OT Risk)", ["(All)"] + dept_list, key="ot_risk_dept")

            try:
                  ot_risk_dept_id = dept_df[dept_df["Department"] == ot_risk_dept]["id"].values[0]
            except IndexError:
                  ot_risk_dept_id = None

            pos_df = refresh(db.Position)
            if ot_risk_dept == "(All)":
                  filtered_positions = pos_df
            else:
                  filtered_positions = pos_df[pos_df["department_id"] == ot_risk_dept_id]

            pos_list = sorted(filtered_positions["name"].dropna().unique())
            ot_risk_pos = st.selectbox("Select Position (OT Risk)", ["(All)"] + pos_list, key="ot_risk_pos")

            # ─────────────── Generate Button ───────────────
            if st.button("📊 Generate OT Risk Report"):

                  # === QUERY ACTUALS ===
                  q = (
                        session.query(
                              db.Actual.emp_id.label("Number"),
                              db.Actual.date.label("Business Date"),
                              (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
                              db.Position.name.label("Position"),
                              db.Department.name.label("Department")
                        )
                        .join(db.Position, db.Actual.position_id == db.Position.id)
                        .join(db.Department, db.Position.department_id == db.Department.id)
                        .filter(db.Actual.date.between(week_start, week_end))
                        .filter(or_(db.Actual.hours != 0, db.Actual.ot_hours != 0))
                  )
                  if ot_risk_dept != "(All)":
                        q = q.filter(db.Department.name == ot_risk_dept)
                  if ot_risk_pos != "(All)":
                        q = q.filter(db.Position.name == ot_risk_pos)

                  raw = pd.DataFrame(q.all(), columns=["Number", "Business Date", "Hours", "Position", "Department"])

                  if raw.empty:
                        st.warning("⚠️ No OT Risk data found. Cannot generate report.")
                  else:
                        raw["Number"] = pd.to_numeric(raw["Number"], errors="coerce").fillna(0).astype(int).astype(str).str.zfill(5)
                        emp_df = refresh(db.Employee).copy()
                        parts = emp_df["name"].astype(str).str.extract(r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)")
                        emp_df["ID"] = parts["ID"].fillna("").astype(str).str.strip().str.zfill(5)
                        emp_df["First Name"] = parts["First_Name"].str.strip()
                        emp_df["Last Name"] = parts["Last_Name"].str.strip()
                        emp_df["match_ID"] = emp_df["ID"].astype(str).str.lstrip("0")
                        raw["match_ID"] = raw["Number"].astype(str).str.lstrip("0")
                        merged = raw.merge(emp_df[["match_ID", "First Name", "Last Name"]], on="match_ID", how="left")

                        agg = merged.groupby(["Number", "First Name", "Last Name"]).agg(
                              total_hours=("Hours", "sum"),
                              days_worked=("Business Date", pd.Series.nunique)
                        ).reset_index()

                        # === QUERY SCHEDULE ===
                        sched_rows = (
                              session.query(db.Employee.name, db.Schedule.day, db.Schedule.shift_type)
                              .join(db.Employee, db.Employee.id == db.Schedule.emp_id)
                              .filter(db.Schedule.day.between(week_start, week_end))
                              .all()
                        )
                        sched_df = pd.DataFrame(sched_rows, columns=["name", "day", "shift_type"])

                        if not sched_df.empty:
                              sched_df["shift_type"] = sched_df["shift_type"].fillna("").astype(str).str.upper().str.strip()
                              sched_df = sched_df[sched_df["shift_type"] != "OFF"]
                              sched_df["Number"] = sched_df["name"].str.extract(r"(\d+)$")[0].fillna("").str.zfill(5)
                              sched_df["day"] = pd.to_datetime(sched_df["day"])
                              merged["Business Date"] = pd.to_datetime(merged["Business Date"])

                              last_worked = (
                                    merged.groupby("Number")["Business Date"]
                                    .max().reset_index().rename(columns={"Business Date": "last_worked"})
                              )

                              sched_df = sched_df.merge(last_worked, on="Number", how="left")
                              sched_df["after_work"] = sched_df["day"] > sched_df["last_worked"]
                              sched_counts = sched_df.groupby("Number")["day"].nunique().reset_index(name="Days Scheduled")
                              sched_future = sched_df[sched_df["after_work"]].copy()
                              days_remaining = sched_future.groupby("Number")["day"].nunique().reset_index(name="Days Remaining")

                              def parse_shift_to_hours(shift_str):
                                    try:
                                          start, end = shift_str.split("-")
                                          start_dt = pd.to_datetime(start, format="%H:%M")
                                          end_dt = pd.to_datetime(end, format="%H:%M")
                                          hours = (end_dt - start_dt).total_seconds() / 3600
                                          if hours < 0:
                                                hours += 24
                                          return max(0, hours - 0.5)
                                    except:
                                          return 0

                              sched_future["shift_hours"] = sched_future["shift_type"].apply(parse_shift_to_hours)
                              future_hours = sched_future.groupby("Number")["shift_hours"].sum().reset_index()
                              future_hours.rename(columns={"shift_hours": "Future Scheduled Hrs"}, inplace=True)

                              agg = agg.merge(sched_counts, how="left", on="Number")
                              agg = agg.merge(days_remaining, how="left", on="Number")
                              agg = agg.merge(future_hours, how="left", on="Number")
                        else:
                              agg["Days Scheduled"] = 0
                              agg["Days Remaining"] = 0
                              agg["Future Scheduled Hrs"] = 0

                        agg["Days Scheduled"] = agg["Days Scheduled"].fillna(0).astype(int)
                        agg["Days Remaining"] = agg["Days Remaining"].fillna(0).astype(int)
                        agg["Future Scheduled Hrs"] = agg["Future Scheduled Hrs"].fillna(0)
                        agg.rename(columns={"days_worked": "Days Worked"}, inplace=True)
                        agg["Total Hrs Worked + Schedule"] = (agg["total_hours"] + agg["Future Scheduled Hrs"]).round(2)

                        def classify_ot_risk(row):
                              if row["Total Hrs Worked + Schedule"] <= 40:
                                    return "No Risk"
                              elif row["Days Remaining"] > 0:
                                    return "At Risk"
                              else:
                                    return "OT"

                        def estimate_risk_percent(row):
                              if row["Total Hrs Worked + Schedule"] <= 40:
                                    return "0%"
                              if row["Days Remaining"] == 0:
                                    return "100%"
                              elif row["Days Remaining"] == 1:
                                    return "80%"
                              elif row["Days Remaining"] == 2:
                                    return "60%"
                              elif row["Days Remaining"] == 3:
                                    return "40%"
                              else:
                                    return "20%"

                        agg["OT Risk"] = agg.apply(classify_ot_risk, axis=1)
                        agg["OT Risk %"] = agg.apply(estimate_risk_percent, axis=1)
                        agg["Projected OT"] = agg["Total Hrs Worked + Schedule"].apply(lambda h: max(round(h - 40, 2), 0))

                        emp_df["ID"] = emp_df["ID"].astype(str)
                        agg["Number"] = agg["Number"].astype(str)
                        emp_df["rate"] = emp_df.get("hourly_rate", 0).fillna(0)
                        agg = agg.merge(emp_df[["ID", "rate"]], left_on="Number", right_on="ID", how="left")
                        agg["rate"] = agg["rate"].fillna(0)
                        agg["OT Cost"] = (agg["Projected OT"] * agg["rate"] * 1.5).round(2)

                        st.subheader("Overtime Risk Report")
                        export_df = agg[[
                              "Number", "First Name", "Last Name", "Days Worked",
                              "Days Scheduled", "Days Remaining", "Total Hrs Worked + Schedule",
                              "OT Risk", "OT Risk %", "Projected OT", "OT Cost"
                        ]].rename(columns={"Total Hrs Worked + Schedule": "Total"})

                        # Export
                        if export_df.empty:
                              st.warning("Report generated, but no OT Risk data was found for the selected filters.")
                        else:
                              if export_format == "Excel":
                                    import io
                                    output = io.BytesIO()
                                    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                                          export_df.to_excel(writer, sheet_name="OT Risk Report", startrow=5, index=False, header=False)
                                          workbook = writer.book
                                          worksheet = writer.sheets["OT Risk Report"]

                                          # Define formats
                                          title_format = workbook.add_format({'bold': True, 'font_size': 14})
                                          label_format = workbook.add_format({'bold': True, 'font_size': 10})
                                          value_format = workbook.add_format({'font_size': 10})

                                          border_fmt = {'border': 1, 'border_color': '#CCCCCC'}
                                          default_fmt = workbook.add_format({**border_fmt})
                                          header_fmt = workbook.add_format({
                                                **border_fmt, 'bold': True, 'bg_color': '#595959', 'font_color': '#FFFFFF', 'align': 'center'
                                          })
                                          first3_fmt = workbook.add_format({
                                                **border_fmt, 'bold': True, 'italic': True
                                          })
                                          last2_fmt = workbook.add_format({
                                                **border_fmt, 'bg_color': '#FEF6F0'
                                          })
                                          green_fmt = workbook.add_format({**border_fmt, 'font_color': '#008000'})
                                          red_fmt = workbook.add_format({**border_fmt, 'font_color': '#FF0000'})
                                          dollar_red = workbook.add_format({**border_fmt, 'num_format': '$#,##0.00', 'font_color': '#FF0000'})
                                          dollar_norm = workbook.add_format({**border_fmt, 'num_format': '$#,##0.00'})
                                          summary_header_fmt = workbook.add_format({
                                                'bold': True, 'bg_color': '#595959', 'font_color': '#FFFFFF', 'border': 1, 'align': 'center'
                                          })
                                          summary_label_fmt = workbook.add_format({
                                                'italic': True, 'border': 1, 'align': 'right'
                                          })
                                          summary_value_fmt = workbook.add_format({
                                                'border': 1
                                          })
                                          dollar_sum_fmt = workbook.add_format({
                                                'border': 1, 'num_format': '$#,##0.00'
                                          })

                                          # Metadata
                                          worksheet.write("A1", "OT Risk Report", title_format)
                                          worksheet.write("A3", "Department:", label_format)
                                          worksheet.write("B3", ot_risk_dept or "(All)", value_format)
                                          worksheet.write("A4", "Position:", label_format)
                                          worksheet.write("B4", ot_risk_pos or "(All)", value_format)
                                          worksheet.write("A5", "Week:", label_format)
                                          worksheet.write("B5", f"{week_start} to {week_end}", value_format)

                                          # Header row
                                          for col_num, col_name in enumerate(export_df.columns):
                                                worksheet.write(5, col_num, col_name, header_fmt)

                                          # Data rows
                                          for row_idx, row in export_df.iterrows():
                                                for col_idx, col_name in enumerate(export_df.columns):
                                                      val = row[col_name]
                                                      if col_idx <= 2:
                                                            fmt = first3_fmt
                                                      elif col_idx >= 9:
                                                            fmt = last2_fmt
                                                      elif col_name == "OT Risk" and val == "OT":
                                                            fmt = red_fmt
                                                      elif col_name == "OT Risk %" and val == "100%":
                                                            fmt = red_fmt
                                                      elif col_name == "OT Risk %" and val == "0%":
                                                            fmt = green_fmt
                                                      elif col_name == "OT Cost":
                                                            try:
                                                                  val = float(val)
                                                                  fmt = dollar_red if val > 0 else dollar_norm
                                                            except:
                                                                  fmt = default_fmt
                                                      else:
                                                            fmt = default_fmt
                                                      worksheet.write(row_idx + 6, col_idx, val, fmt)

                                          # Summary block (outside table)
                                          summary_row = 1
                                          summary_col = len(export_df.columns) - 2

                                          worksheet.merge_range(
                                                summary_row,
                                                summary_col,
                                                summary_row,
                                                summary_col + 1,
                                                "Total",
                                                summary_header_fmt
                                          )
                                          worksheet.write(summary_row + 1, summary_col, "Projected OT", summary_label_fmt)
                                          worksheet.write(summary_row + 1, summary_col + 1, export_df["Projected OT"].sum(), summary_value_fmt)
                                          worksheet.write(summary_row + 2, summary_col, "OT Cost", summary_label_fmt)
                                          worksheet.write(summary_row + 2, summary_col + 1, export_df["OT Cost"].sum(), dollar_sum_fmt)

                                          # Auto column width
                                          for i, col in enumerate(export_df.columns):
                                                max_width = max(len(str(col)), export_df[col].astype(str).str.len().max())
                                                worksheet.set_column(i, i, max_width + 2)

                                    st.download_button(
                                          "⬇️ Download Excel",
                                          data=output.getvalue(),
                                          file_name=f"OT_Risk_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.xlsx",
                                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )

                              elif export_format == "PDF":
                                    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                                    from reportlab.lib.pagesizes import landscape, letter
                                    from reportlab.lib.styles import getSampleStyleSheet
                                    from reportlab.lib import colors

                                    buffer = io.BytesIO()
                                    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
                                    styles = getSampleStyleSheet()
                                    elements = []

                                    # ───── Title + Metadata ─────
                                    elements.append(Paragraph("OT Risk Report", styles["Heading2"]))
                                    elements.append(Spacer(1, 6))
                                    elements.append(Paragraph(f"<b>Department:</b> {ot_risk_dept or '(All)'}", styles["Normal"]))
                                    elements.append(Paragraph(f"<b>Position:</b> {ot_risk_pos or '(All)'}", styles["Normal"]))
                                    elements.append(Paragraph(f"<b>Week:</b> {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}", styles["Normal"]))
                                    elements.append(Spacer(1, 12))

                                    # ───── Table Data ─────
                                    data = [export_df.columns.tolist()] + export_df.values.tolist()

                                    table = Table(data)
                                    table.setStyle(TableStyle([
                                          ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#595959")),
                                          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                          ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                          ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                          ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                                          ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                                    ]))

                                    elements.append(table)
                                    elements.append(Spacer(1, 12))

                                    # ───── Summary Table ─────
                                    summary_data = [
                                          ["Total", ""],
                                          ["Projected OT", f"{export_df['Projected OT'].sum():.2f}"],
                                          ["OT Cost", f"${export_df['OT Cost'].sum():,.2f}"],
                                    ]

                                    summary_table = Table(summary_data, hAlign='RIGHT')
                                    summary_table.setStyle(TableStyle([
                                          ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#595959")),
                                          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                          ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                          ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                          ("FONTNAME", (0, 1), (0, -1), "Helvetica-Oblique"),
                                          ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                                          ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FEF6F0")),
                                    ]))

                                    elements.append(summary_table)

                                    doc.build(elements)
                                    st.download_button(
                                          "⬇️ Download PDF",
                                          data=buffer.getvalue(),
                                          file_name=f"OT_Risk_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.pdf",
                                          mime="application/pdf"
                                    )

      if report_type == "Forecast Variance":
            st.markdown("### 🔍 Forecast Variance Filters")

            if "forecast_var_date" not in st.session_state:
                  st.session_state.forecast_var_date = date.today()

            sel_date = st.date_input("Select any date in the week", value=st.session_state.forecast_var_date, key="forecast_var_date")
            week_start = sel_date + relativedelta(weekday=MO(-1))
            week_end   = week_start + timedelta(days=6)

            st.markdown(f"📅 **Selected Week:** {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}")

            generate_forecast_var = st.button("📊 Generate Forecast Variance Report")

            if generate_forecast_var:
                  def pull_week_kpi_totals(Model, label):
                        rows = (
                              session.query(Model)
                                     .filter(Model.date.between(week_start, week_end))
                                     .all()
                        )
                        data = defaultdict(float)
                        for r in rows:
                              data[r.kpi] += r.value
                        return pd.DataFrame([(k, v) for k, v in data.items()], columns=["KPI", label])

                  df_forecast = pull_week_kpi_totals(db.RoomForecast, "Forecast")
                  df_actual   = pull_week_kpi_totals(db.RoomActual, "Actual")
                  df_otb      = pull_week_kpi_totals(db.RoomOTBPickup, "OTB + Pickup")
                  # If no data found for the selected week, show warning and stop
                  if df_forecast.empty and df_actual.empty and df_otb.empty:
                        st.warning("⚠️ No forecast, actual, or OTB + pickup data found for the selected week.")
                        st.stop()

                  merged = df_forecast.merge(df_actual, on="KPI", how="outer")\
                                      .merge(df_otb, on="KPI", how="outer").fillna(0)

                  merged = merged[["KPI", "Actual", "Forecast", "OTB + Pickup"]]
                  merged["Δ Actual - Forecast"] = merged["Actual"] - merged["Forecast"]
                  merged["Δ OTB - Forecast"]    = merged["OTB + Pickup"] - merged["Forecast"]

                  def add_arrow(val):
                        if val > 0:
                              return f"▲ {val}"
                        elif val < 0:
                              return f"▼ {abs(val)}"
                        else:
                              return "0"

                  merged["Δ Actual - Forecast"] = merged["Δ Actual - Forecast"].apply(add_arrow)
                  merged["Δ OTB - Forecast"]    = merged["Δ OTB - Forecast"].apply(add_arrow)

                  export_df = merged.copy()

                  if export_format == "Excel":
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                              export_df.to_excel(writer, sheet_name="Forecast Variance", index=False, startrow=5, header=False)
                              workbook = writer.book
                              worksheet = writer.sheets["Forecast Variance"]

                              title_fmt = workbook.add_format({'bold': True, 'font_size': 14})
                              label_fmt = workbook.add_format({'bold': True})
                              value_fmt = workbook.add_format({})
                              kpi_fmt   = workbook.add_format({'bold': True, 'italic': True, 'border': 1})
                              header_fmt = workbook.add_format({
                                    'bold': True, 'bg_color': '#595959', 'font_color': '#FFFFFF', 'border': 1, 'align': 'center'
                              })
                              default_fmt = workbook.add_format({'border': 1})
                              variance_fmt = workbook.add_format({'bg_color': '#FEF6F0', 'border': 1, 'align': 'right'})

                              worksheet.write("A1", "Forecast Variance Report", title_fmt)
                              worksheet.write("A3", "Week:", label_fmt)
                              worksheet.write("B3", f"{week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}", value_fmt)

                              for col_idx, col in enumerate(export_df.columns):
                                    worksheet.write(5, col_idx, col, header_fmt)

                              for row_idx, row in export_df.iterrows():
                                    for col_idx, col in enumerate(export_df.columns):
                                          val = row[col]
                                          if col == "KPI":
                                                fmt = kpi_fmt
                                          elif "Δ" in col:
                                                fmt = variance_fmt
                                          else:
                                                fmt = default_fmt
                                          worksheet.write(row_idx + 6, col_idx, val, fmt)

                              worksheet.set_column(0, len(export_df.columns) - 1, 18)

                        st.download_button(
                              "⬇️ Download Excel",
                              data=output.getvalue(),
                              file_name=f"Forecast_Variance_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                  elif export_format == "PDF":
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib.pagesizes import landscape, letter
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors

                        buffer = io.BytesIO()
                        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
                        styles = getSampleStyleSheet()
                        elements = []

                        # ───── Title + Metadata ─────
                        elements.append(Paragraph("Forecast Variance Report", styles["Heading2"]))
                        elements.append(Spacer(1, 6))
                        elements.append(Paragraph(f"<b>Week:</b> {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}", styles["Normal"]))
                        elements.append(Spacer(1, 12))

                        # ───── Table Data ─────
                        data = [export_df.columns.tolist()] + export_df.values.tolist()

                        table = Table(data)
                        table.setStyle(TableStyle([
                              ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#595959")),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                              ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                        ]))

                        # Apply KPI column formatting + Δ column background
                        for col_idx, col_name in enumerate(export_df.columns):
                              if col_name == "KPI":
                                    table.setStyle([("FONTNAME", (col_idx, 1), (col_idx, -1), "Helvetica-Oblique"),
                                                    ("FONTNAME", (col_idx, 1), (col_idx, -1), "Helvetica-Bold")])
                              if "Δ" in col_name:
                                    table.setStyle([("BACKGROUND", (col_idx, 1), (col_idx, -1), colors.HexColor("#FEF6F0"))])

                        elements.append(table)

                        # ───── Export ─────
                        doc.build(elements)
                        st.download_button(
                              "⬇️ Download PDF",
                              data=buffer.getvalue(),
                              file_name=f"Forecast_Variance_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.pdf",
                              mime="application/pdf"
                        )

                  elif export_format == "CSV":
                        csv_data = export_df.to_csv(index=False).encode("utf-8")
                        st.download_button(
                              "⬇️ Download CSV",
                              data=csv_data,
                              file_name=f"Forecast_Variance_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.csv",
                              mime="text/csv"
                        )
      if report_type == "Department Schedule":
            st.markdown("### 🗓️ Department Schedule Filters")

            import os
            from datetime import date, timedelta
            from dateutil.relativedelta import relativedelta, MO

            fmt_day = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"

            if "dept_sched_date" not in st.session_state:
                  st.session_state.dept_sched_date = date.today()

            sel_date = st.date_input("📆 Select any date in the week", value=st.session_state.dept_sched_date, key="dept_sched_date")
            week_start = sel_date + relativedelta(weekday=MO(-1))
            week_end   = week_start + timedelta(days=6)
            week_dates = [week_start + timedelta(i) for i in range(7)]
            day_cols   = [d.strftime(fmt_day) for d in week_dates]

            st.session_state["dept_schedule_week_start"] = week_start
            st.session_state["dept_schedule_week_end"] = week_end
            st.session_state["dept_schedule_day_cols"] = day_cols

            st.markdown(f"📅 **Selected Week:** {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}")

            emp_df = refresh(db.Employee)
            dept_list = sorted(emp_df["department"].dropna().unique())
            sel_dept = st.selectbox("🏢 Department*", dept_list, key="sched_dept_filter")

            pos_opts = emp_df[emp_df["department"] == sel_dept]["role"].dropna().unique()
            if len(pos_opts) == 0:
                  st.warning("⚠️ No positions found under the selected department.")
                  st.stop()

            sel_pos = st.selectbox("🧑‍💼 Position*", sorted(pos_opts), key="sched_pos_filter")
            generate_schedule = st.button("📊 Generate Schedule Report")

            if generate_schedule:
                  if not sel_dept or not sel_pos:
                        st.warning("Please select both a department and a position.")
                        st.stop()

                  emp_sub = emp_df[(emp_df["department"] == sel_dept) & (emp_df["role"] == sel_pos)]

                  if emp_sub.empty:
                        st.warning("⚠️ No matching employees for selected filters.")
                        st.stop()

                  ids     = emp_sub["name"].str.extract(r"(\d+)$")[0].fillna("")
                  firsts  = emp_sub["name"].str.extract(r",\s*([^\d]+)")[0].str.strip()
                  lasts   = emp_sub["name"].str.extract(r"^\s*([^,]+)")[0].str.strip()

                  sched_df = pd.DataFrame({
                        "ID": ids,
                        "First Name": firsts,
                        "Last Name": lasts,
                        "emp_id": emp_sub["id"]
                  })
                  for col in day_cols:
                        sched_df[col] = ""

                  sched_rows = session.query(db.Schedule).filter(
                        db.Schedule.emp_id.in_(emp_sub["id"]),
                        db.Schedule.day.in_(week_dates)
                  ).all()

                  if not sched_rows:
                        st.warning("⚠️ No schedule data found for this position and week.")
                        st.stop()

                  for r in sched_rows:
                        col = r.day.strftime(fmt_day)
                        idx = sched_df.index[sched_df["emp_id"] == r.emp_id]
                        if not idx.empty:
                              sched_df.at[idx[0], col] = r.shift_type

                  sched_df.drop(columns=["emp_id"], inplace=True)
                  st.session_state["dept_schedule_df"] = sched_df
            if "dept_schedule_df" in st.session_state and generate_schedule:
                  df = st.session_state["dept_schedule_df"]
                  week_start = st.session_state["dept_schedule_week_start"]
                  week_end   = st.session_state["dept_schedule_week_end"]
                  dept_name  = sel_dept
                  report_title = f"Department Schedule Report – {dept_name} – Week of {week_start:%B %d}–{week_end:%d}"

                  if export_format == "Excel":
                        import io
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                        from openpyxl.utils import get_column_letter

                        buffer = io.BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = f"Schedule – {sel_pos}"

                        # ─────────── Styles
                        bold_font = Font(bold=True)
                        center_align = Alignment(horizontal="center", vertical="center")
                        header_fill = PatternFill("solid", fgColor="444444")
                        header_font = Font(color="FFFFFF", bold=True)
                        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                        top=Side(style='thin'), bottom=Side(style='thin'))
                        off_fill = PatternFill("solid", fgColor="FEBFBA")
                        total_fill = PatternFill("solid", fgColor="FEF6F0")
                        red_font = Font(color="FF0000")
                        bold_italic = Font(bold=True, italic=True)

                        # ─────────── Title + Metadata
                        ws["A1"] = "Department Schedule Report"
                        ws["A1"].font = Font(bold=True, size=14)
                        ws.merge_cells("A1:D1")

                        ws["A3"] = "Department:"
                        ws["B3"] = sel_dept
                        ws["A4"] = "Position:"
                        ws["B4"] = sel_pos
                        ws["A5"] = "Week:"
                        ws["B5"] = f"{week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}"

                        # ─────────── Total Column Calculation
                        df["Total"] = df.iloc[:, 3:].apply(
                              lambda row: sum(
                                    (
                                          float(end.split(":")[0]) + float(end.split(":")[1]) / 60
                                          - float(start.split(":")[0]) - float(start.split(":")[1]) / 60
                                    ) if isinstance(x, str) and "-" in x and ":" in x
                                    and len(x.split("-")) == 2
                                    and (start := x.split("-")[0]) and (end := x.split("-")[1])
                                    else 0
                                    for x in row
                              ),
                              axis=1
                        ).round(2)

                        # ─────────── Data Table
                        start_row = 7
                        for i, col in enumerate(df.columns, start=1):
                              cell = ws.cell(row=start_row, column=i, value=col)
                              cell.font = header_font
                              cell.fill = header_fill
                              cell.alignment = center_align
                              cell.border = border

                        for r_idx, row in enumerate(df.values.tolist(), start=start_row + 1):
                              for c_idx, val in enumerate(row, start=1):
                                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                                    cell.alignment = center_align
                                    cell.border = border

                                    if c_idx <= 3:
                                          cell.font = bold_italic

                                    if isinstance(val, str) and val.strip().upper() == "OFF":
                                          cell.fill = off_fill

                                    if df.columns[c_idx - 1] == "Total":
                                          cell.fill = total_fill
                                          cell.number_format = "0.00"
                                          if isinstance(val, (int, float)) and val > 40:
                                                cell.font = red_font

                        # ─────────── Auto Column Widths
                        for col_idx in range(1, df.shape[1] + 1):
                              col_letter = get_column_letter(col_idx)
                              values = [ws[f"{col_letter}{r}"].value for r in range(start_row, start_row + df.shape[0] + 1)]
                              max_len = max(len(str(v) if v is not None else "") for v in values)
                              ws.column_dimensions[col_letter].width = max_len + 2

                        wb.save(buffer)
                        st.download_button(
                              "📥 Download Excel",
                              buffer.getvalue(),
                              file_name=f"Dept_Schedule_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.xlsx"
                        )

                  elif export_format == "PDF":
                        import io
                        from reportlab.lib.pagesizes import landscape, letter
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors

                        buffer = io.BytesIO()
                        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
                        styles = getSampleStyleSheet()
                        elements = []

                        # ─────────── Title + Metadata
                        elements.append(Paragraph("Department Schedule Report", styles["Heading2"]))
                        elements.append(Spacer(1, 6))
                        elements.append(Paragraph(f"<b>Department:</b> {sel_dept}", styles["Normal"]))
                        elements.append(Paragraph(f"<b>Position:</b> {sel_pos}", styles["Normal"]))
                        elements.append(Paragraph(f"<b>Week:</b> {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}", styles["Normal"]))
                        elements.append(Spacer(1, 12))

                        # ─────────── Total Column Calculation
                        df["Total"] = df.iloc[:, 3:].apply(
                              lambda row: sum(
                                    (
                                          float(end.split(":")[0]) + float(end.split(":")[1]) / 60
                                          - float(start.split(":")[0]) - float(start.split(":")[1]) / 60
                                    ) if isinstance(x, str) and "-" in x and ":" in x
                                    and len(x.split("-")) == 2
                                    and (start := x.split("-")[0]) and (end := x.split("-")[1])
                                    else 0
                                    for x in row
                              ),
                              axis=1
                        ).round(2)

                        # ─────────── Table Setup
                        data = [df.columns.tolist()] + df.values.tolist()
                        table = Table(data, repeatRows=1)

                        # ─────────── Style Definitions
                        style = TableStyle([
                              ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#444444")),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                              ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ])

                        # ─────────── Cell-level Styling
                        for r_idx, row in enumerate(data[1:], start=1):
                              for c_idx, val in enumerate(row):
                                    if c_idx <= 2:
                                          style.add("FONTNAME", (c_idx, r_idx), (c_idx, r_idx), "Helvetica-BoldOblique")
                                    if isinstance(val, str) and val.strip().upper() == "OFF":
                                          style.add("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#FEBFBA"))
                                    if df.columns[c_idx] == "Total":
                                          style.add("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#FEF6F0"))
                                          if isinstance(val, (int, float)) and val > 40:
                                                style.add("TEXTCOLOR", (c_idx, r_idx), (c_idx, r_idx), colors.red)

                        table.setStyle(style)
                        elements.append(table)

                        doc.build(elements)
                        st.download_button(
                              "📥 Download PDF",
                              buffer.getvalue(),
                              file_name=f"Dept_Schedule_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.pdf"
                        )
                  elif export_format == "CSV":
                        csv = df.to_csv(index=False)
                        st.download_button(
                              "📥 Download CSV",
                              csv,
                              file_name=f"Dept_Schedule_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.csv"
                        )

      elif report_type == "Productivity Index":
            st.markdown("### 📈 Productivity Index Report")

            if "prod_index_start" not in st.session_state:
                  st.session_state.prod_index_start = date.today()
            if "prod_index_end" not in st.session_state:
                  st.session_state.prod_index_end = date.today()

            col1, col2 = st.columns(2)
            with col1:
                  sel_start = st.date_input("From Date", value=st.session_state.prod_index_start, key="prod_index_start")
            with col2:
                  sel_end = st.date_input("To Date", value=st.session_state.prod_index_end, key="prod_index_end")
            week_start = sel_start
            week_end   = sel_end

            dept_df = refresh(db.Department).rename(columns={"id": "dept_id", "name": "dept"})
            pos_df  = refresh(db.Position).rename(columns={"id": "position_id", "name": "position"})

            sel_dept = st.selectbox("Select Department", sorted(dept_df["dept"].dropna().unique()), key="prod_dept")
            filtered_pos = pos_df.merge(dept_df, left_on="department_id", right_on="dept_id")
            pos_options = ["All Positions"] + sorted(filtered_pos[filtered_pos["dept"] == sel_dept]["position"].unique().tolist())
            sel_pos = st.selectbox("Select Position", pos_options, key="prod_pos")

            generate_btn = st.button("📊 Generate Productivity Report")

            if generate_btn:
                  std_df = refresh(db.LaborStandard)
                  std_df = std_df.merge(pos_df[["position_id", "position", "department_id"]], on="position_id", how="left")
                  std_df = std_df.merge(dept_df[["dept_id", "dept"]], left_on="department_id", right_on="dept_id", how="left")
                  std_df = std_df[std_df["dept"] == sel_dept]
                  if sel_pos != "All Positions":
                        std_df = std_df[std_df["position"] == sel_pos]

                  ah_df = refresh(db.Actual)
                  ah_df = ah_df[
                        (ah_df["source"].isin(["manual", "contract"])) &
                        (ah_df["date"].between(week_start, week_end))
                  ]
                  ah_df["total_hours"] = ah_df[["hours", "ot_hours"]].sum(axis=1)
                  hours_summary = (
                        ah_df.groupby("position_id")["total_hours"]
                        .sum()
                        .reset_index()
                        .rename(columns={"total_hours": "actual_hours"})
                  )
                  hours_summary = hours_summary.merge(pos_df[["position_id", "position"]], on="position_id", how="left")

                  kpi_df = refresh(db.RoomActual)
                  kpi_df = kpi_df[kpi_df["date"].between(week_start, week_end)]
                  kpi_summary = (
                        kpi_df.groupby("kpi")["value"]
                        .sum()
                        .reset_index()
                        .rename(columns={"value": "output"})
                  )

                  final_rows = []
                  for _, row in std_df.iterrows():
                        position = row["position"]
                        metric = row["metric"]
                        raw_standard = row["standard"]
                        standard = 8 / raw_standard if raw_standard else None

                        actual_hours = hours_summary.loc[
                              hours_summary["position"] == position, "actual_hours"
                        ].sum()

                        output = kpi_summary.loc[
                              kpi_summary["kpi"] == metric, "output"
                        ].sum()

                        if actual_hours > 0 and output > 0 and standard is not None:
                              productivity = actual_hours / output
                              variance = standard - productivity
                              arrow = "▲" if variance > 0 else ("▼" if variance < 0 else "")
                              final_rows.append({
                                    "Position": position,
                                    "KPI": metric,
                                    "Output": round(output, 2),
                                    "Hours": round(actual_hours, 2),
                                    "Productivity (hrs/unit)": round(productivity, 2),
                                    "Standard (hrs/unit)": round(standard, 2),
                                    "Variance": f"{round(variance, 2)} {arrow}"
                              })

                  final_df = pd.DataFrame(final_rows)

                  if final_df.empty:
                        st.warning("⚠️ No productivity data available for the selected period.")
                  else:
                        total_output = final_df["Output"].sum()
                        total_hours = final_df["Hours"].sum()
                        weighted_productivity = total_hours / total_output if total_output else 0
                        avg_standard = final_df["Standard (hrs/unit)"].astype(float).mean()
                        variance = avg_standard - weighted_productivity
                        arrow = "▲" if variance > 0 else ("▼" if variance < 0 else "")

                        total_row = {
                              "Position": "TOTAL",
                              "KPI": "",
                              "Output": round(total_output, 2),
                              "Hours": round(total_hours, 2),
                              "Productivity (hrs/unit)": round(weighted_productivity, 2),
                              "Standard (hrs/unit)": round(avg_standard, 2),
                              "Variance": f"{round(variance, 2)} {arrow}"
                        }
                        final_df = pd.concat([final_df, pd.DataFrame([total_row])], ignore_index=True)

                        if export_format == "Excel":
                              buffer = io.BytesIO()
                              with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
                                    workbook = writer.book
                                    worksheet_name = "Productivity"
                                    final_df.to_excel(writer, index=False, sheet_name=worksheet_name, startrow=6, header=False)
                                    worksheet = writer.sheets[worksheet_name]

                                    # Title + metadata
                                    title_fmt = workbook.add_format({"bold": True, "font_size": 14})
                                    meta_fmt = workbook.add_format({"italic": True, "font_size": 10})
                                    worksheet.write("A1", "Productivity Index Report", title_fmt)
                                    worksheet.write("A3", f"Department: {sel_dept}", meta_fmt)
                                    worksheet.write("A4", f"Position: {sel_pos}", meta_fmt)
                                    worksheet.write("A5", f"Period: {week_start:%b %d, %Y} – {week_end:%b %d, %Y}", meta_fmt)

                                    # Header style
                                    header_fmt = workbook.add_format({
                                          "bold": True, "font_color": "white", "bg_color": "#595959", "border": 1, "align": "center"
                                    })
                                    for col_num, value in enumerate(final_df.columns.values):
                                          worksheet.write(6, col_num, value, header_fmt)

                                    # Set column widths + center alignment
                                    center_fmt = workbook.add_format({"align": "center"})
                                    worksheet.set_column("A:G", 20, center_fmt)

                                    # Border format for data rows (excluding total row)
                                    border_fmt = workbook.add_format({"align": "center", "border": 1})
                                    for row_idx, row in final_df[:-1].iterrows():
                                          for col_idx, value in enumerate(row):
                                                worksheet.write(row_idx + 7, col_idx, value, border_fmt)

                                    # Highlight TOTAL row
                                    total_row_idx = len(final_df) + 6
                                    total_fmt = workbook.add_format({
                                          "bold": True,
                                          "italic": True,
                                          "bg_color": "#FEF6F0",
                                          "align": "center",
                                          "border": 1
                                    })
                                    for col_idx, value in enumerate(final_df.iloc[-1]):
                                          worksheet.write(total_row_idx, col_idx, value, total_fmt)

                              file_name = f"Productivity_Index_{week_start:%Y%m%d}_{week_end:%Y%m%d}.xlsx"
                              st.download_button("📥 Download Excel", buffer.getvalue(), file_name=file_name)

                        elif export_format == "PDF":
                              import io
                              from reportlab.lib.pagesizes import landscape, letter
                              from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                              from reportlab.lib.styles import getSampleStyleSheet
                              from reportlab.lib import colors

                              buffer = io.BytesIO()
                              doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
                              styles = getSampleStyleSheet()
                              elements = []

                              # ───────── Title + Metadata
                              elements.append(Paragraph("Productivity Index Report", styles["Heading2"]))
                              elements.append(Spacer(1, 6))
                              elements.append(Paragraph(f"<b>Department:</b> {sel_dept}", styles["Normal"]))
                              elements.append(Paragraph(f"<b>Position:</b> {sel_pos}", styles["Normal"]))
                              elements.append(Paragraph(f"<b>Period:</b> {week_start:%b %d, %Y} – {week_end:%b %d, %Y}", styles["Normal"]))
                              elements.append(Spacer(1, 12))

                              # ───────── Table Setup
                              data = [final_df.columns.tolist()] + final_df.astype(str).values.tolist()
                              table = Table(data, repeatRows=1)

                              # ───────── Style
                              style = TableStyle([
                                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#595959")),
                                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                              ])

                              # ───────── Row Styling
                              for r_idx, row in enumerate(data[1:], start=1):
                                    is_total = row[0] == "TOTAL"
                                    for c_idx, val in enumerate(row):
                                          if c_idx == 0:  # First column bold+italic
                                                style.add("FONTNAME", (c_idx, r_idx), (c_idx, r_idx), "Helvetica-BoldOblique")
                                          if is_total:
                                                style.add("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#FEF6F0"))

                              table.setStyle(style)
                              elements.append(table)

                              doc.build(elements)
                              st.download_button(
                                    "📥 Download PDF",
                                    buffer.getvalue(),
                                    file_name=f"Productivity_Index_{week_start:%Y%m%d}_{week_end:%Y%m%d}.pdf"
                              )

                        elif export_format == "CSV":
                              # Remove arrows from Variance column for CSV export
                              clean_df = final_df.copy()
                              if "Variance" in clean_df.columns:
                                    clean_df["Variance"] = clean_df["Variance"].astype(str).str.replace("▲", "", regex=False).str.replace("▼", "", regex=False).str.strip()

                              csv_buffer = io.StringIO()
                              clean_df.to_csv(csv_buffer, index=False)

                              file_name = f"Productivity_Index_{week_start:%Y%m%d}_{week_end:%Y%m%d}.csv"
                              st.download_button(
                                    "📥 Download CSV",
                                    data=csv_buffer.getvalue().encode("utf-8"),
                                    file_name=file_name,
                                    mime="text/csv"
                              )

      elif report_type == "Labor Standards":
            st.markdown("### 📏 Labor Standards Report")

            generate_std = st.button("📊 Generate Report")

            if generate_std:
                  import pandas as pd

                  std_df = pd.read_sql("""
                        SELECT d.name AS Department, 
                               p.name AS Position, 
                               s.metric AS Metric, 
                               s.standard AS Standard, 
                               s.unit AS Unit
                        FROM labor_standards s
                        JOIN positions p ON s.position_id = p.id
                        JOIN departments d ON p.department_id = d.id
                        ORDER BY d.name, p.name, s.metric
                  """, con=ENGINE)

                  if std_df.empty:
                        st.warning("⚠️ No labor standards found in the database.")
                  else:
                        if export_format == "Excel":
                              from io import BytesIO
                              import xlsxwriter

                              output = BytesIO()
                              with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                                    workbook  = writer.book
                                    worksheet = workbook.add_worksheet("Labor Standards")
                                    writer.sheets["Labor Standards"] = worksheet

                                    title_format = workbook.add_format({'bold': True, 'font_size': 14})
                                    header_format = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#404040', 'align': 'center', 'valign': 'vcenter', 'border': 1})
                                    center_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
                                    bold_italic_center = workbook.add_format({'bold': True, 'italic': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
                                    pink_fill_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FEF6F0'})

                                    worksheet.write("A1", "Labor Standards Report", title_format)
                                    worksheet.set_row(0, 24)
                                    worksheet.set_column("A:E", 20)

                                    for col_num, value in enumerate(std_df.columns.values):
                                          worksheet.write(2, col_num, value, header_format)

                                    for row_num, row in enumerate(std_df.values):
                                          for col_num, cell in enumerate(row):
                                                if col_num in [0, 1]:
                                                      worksheet.write(row_num + 3, col_num, cell, bold_italic_center)
                                                elif col_num in [3, 4]:
                                                      worksheet.write(row_num + 3, col_num, cell, pink_fill_format)
                                                else:
                                                      worksheet.write(row_num + 3, col_num, cell, center_format)

                              st.download_button(
                                    label="📥 Download Excel",
                                    data=output.getvalue(),
                                    file_name="Labor_Standards_Report.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                              )

                        elif export_format == "PDF":
                              from io import BytesIO
                              from reportlab.lib.pagesizes import landscape, letter
                              from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                              from reportlab.lib.styles import getSampleStyleSheet
                              from reportlab.lib import colors

                              buffer = BytesIO()
                              doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
                              styles = getSampleStyleSheet()
                              elements = []

                              # Title
                              elements.append(Paragraph("Labor Standards Report", styles["Heading2"]))
                              elements.append(Spacer(1, 12))

                              data = [list(std_df.columns)] + std_df.values.tolist()

                              # Style table
                              table = Table(data)
                              table.setStyle(TableStyle([
                                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#404040")),
                                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                                    ('FONTNAME', (0, 1), (1, -1), 'Helvetica-BoldOblique'),
                                    ('BACKGROUND', (3, 1), (4, -1), colors.HexColor("#FEF6F0")),
                              ]))
                              elements.append(table)

                              doc.build(elements)

                              st.download_button(
                                    label="📥 Download PDF",
                                    data=buffer.getvalue(),
                                    file_name="Labor_Standards_Report.pdf",
                                    mime="application/pdf"
                              )

                        elif export_format == "CSV":
                              csv_data = std_df.to_csv(index=False).encode("utf-8")
                              st.download_button(
                                    label="📥 Download CSV",
                                    data=csv_data,
                                    file_name="Labor_Standards_Report.csv",
                                    mime="text/csv"
                              )

      elif report_type == "Labor Variance":
            st.markdown("### 📘 Labor Variance Report")

            from datetime import date, timedelta
            from dateutil.relativedelta import relativedelta, MO
            import pandas as pd
            import io

            if "labor_var_date" not in st.session_state:
                  st.session_state.labor_var_date = date.today()

            sel_date = st.date_input("Select any date in the week", value=st.session_state.labor_var_date, key="labor_var_date")
            week_start = sel_date + relativedelta(weekday=MO(-1))
            week_end   = week_start + timedelta(days=6)

            st.markdown(f"📅 **Selected Week:** {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}")

            emp_df = refresh(db.Employee)
            dept_list = sorted(emp_df["department"].dropna().unique())
            sel_dept = st.selectbox("🏢 Department", dept_list, key="labor_var_dept")

            pos_list = emp_df[emp_df["department"] == sel_dept]["role"].dropna().unique()
            sel_pos = st.selectbox("🧑‍💼 Position", ["All"] + sorted(pos_list), key="labor_var_pos")
            # ───── Pre-check for Data Availability ─────
            schedule_df = refresh(db.Schedule)
            actual_df   = refresh(db.Actual)
            room_df     = refresh(db.RoomActual)

            has_schedule = not schedule_df[(schedule_df["day"] >= week_start) & (schedule_df["day"] <= week_end)].empty
            has_actual   = not actual_df[(actual_df["date"] >= week_start) & (actual_df["date"] <= week_end)].empty
            has_room     = not room_df[(room_df["date"] >= week_start) & (room_df["date"] <= week_end)].empty



            generate_labor_var = st.button("📊 Generate Labor Variance Report")
            if generate_labor_var:
                  schedule_df     = refresh(db.Schedule)
                  actual_df       = refresh(db.Actual)
                  room_actual_df  = refresh(db.RoomActual)
                  std_df          = refresh(db.LaborStandard)
                  pos_df          = refresh(db.Position).rename(columns={"id": "position_id", "name": "position"})

                  has_schedule = not schedule_df[(schedule_df["day"] >= week_start) & (schedule_df["day"] <= week_end)].empty
                  has_actual   = not actual_df[(actual_df["date"] >= week_start) & (actual_df["date"] <= week_end)].empty
                  has_room     = not room_actual_df[(room_actual_df["date"] >= week_start) & (room_actual_df["date"] <= week_end)].empty

                  if not (has_schedule or has_actual or has_room):
                        st.warning("⚠️ No data available for the selected week. Labor Variance Report cannot be generated.")
                        st.session_state.labor_variance_ready = False
                        st.stop()

                  emp_filtered = emp_df[emp_df["department"] == sel_dept]
                  if sel_pos != "All":
                        emp_filtered = emp_filtered[emp_filtered["role"] == sel_pos]

                  pos_names = emp_filtered["role"].dropna().unique()
                  pos_match = pos_df[pos_df["position"].isin(pos_names)]

                  if pos_match.empty:
                        st.warning("⚠️ No positions found matching employee roles.")
                        st.session_state.labor_variance_ready = False
                        st.stop()

                  ah_df = actual_df[
                        (actual_df["source"].isin(["manual", "contract"])) &
                        (actual_df["date"].between(week_start, week_end))
                  ]
                  ah_df["total_hours"] = ah_df[["hours", "ot_hours"]].sum(axis=1)
                  actual_summary = (
                        ah_df.groupby("position_id")["total_hours"]
                        .sum()
                        .reset_index()
                        .rename(columns={"total_hours": "actual_hours"})
                  )
                  actual_summary = actual_summary.merge(pos_df[["position_id", "position"]], on="position_id", how="left")

                  results = []
                  seen_positions = set()

                  for _, row in pos_match.iterrows():
                        pos_name = row["position"]
                        pos_id   = row["position_id"]

                        if pos_name in seen_positions:
                              continue
                        seen_positions.add(pos_name)

                        emp_ids = emp_df[
                              (emp_df["department"] == sel_dept) &
                              (emp_df["role"] == pos_name)
                        ]["id"].tolist()

                        sched_rows = schedule_df[
                              (schedule_df["emp_id"].isin(emp_ids)) &
                              (schedule_df["day"] >= week_start) &
                              (schedule_df["day"] <= week_end)
                        ]
                        sched_hours = sched_rows["shift_type"].apply(lambda x: 0 if x == "OFF" else (
                              pd.to_datetime(x.split("-")[1]) - pd.to_datetime(x.split("-")[0])
                        ).seconds / 3600).sum()

                        actual_hours = actual_summary.loc[
                              actual_summary["position"] == pos_name, "actual_hours"
                        ].sum()

                        std_pos = std_df[std_df["position_id"] == pos_id]
                        proj_hours_total = 0
                        for _, std_row in std_pos.iterrows():
                              metric = std_row["metric"]
                              standard = std_row["standard"]
                              if not standard or standard == 0:
                                    continue

                              actual_output = room_actual_df[
                                    (room_actual_df["kpi"] == metric) &
                                    (room_actual_df["date"] >= week_start) &
                                    (room_actual_df["date"] <= week_end)
                              ]["value"].sum()

                              proj_hours_total += (actual_output / standard) * 8

                        projected_hours = proj_hours_total
                        variance = actual_hours - projected_hours
                        variance_pct = (variance / projected_hours * 100) if projected_hours else 0

                        results.append({
                              "Position": pos_name,
                              "Scheduled Hours": round(sched_hours, 1),
                              "Actual Hours": round(actual_hours, 1),
                              "Projected Hours": round(projected_hours, 1),
                              "Variance": round(variance, 1),
                              "Variance %": f"{'▲' if variance > 0 else '▼' if variance < 0 else ''} {abs(variance_pct):.2f}%" if projected_hours else "–"
                        })

                  report_df = pd.DataFrame(results)

                  # ───── Add Total Row After Deduplication ─────
                  total_row = {
                        "Position": "TOTAL",
                        "Scheduled Hours": round(report_df["Scheduled Hours"].sum(), 1),
                        "Actual Hours": round(report_df["Actual Hours"].sum(), 1),
                        "Projected Hours": round(report_df["Projected Hours"].sum(), 1),
                        "Variance": round(report_df["Variance"].sum(), 1)
                  }

                  if total_row["Projected Hours"]:
                        total_var_pct = (total_row["Variance"] / total_row["Projected Hours"]) * 100
                        total_row["Variance %"] = f"{'▲' if total_var_pct > 0 else '▼' if total_var_pct < 0 else ''} {abs(total_var_pct):.2f}%"
                  else:
                        total_row["Variance %"] = "–"

                  report_df.loc[len(report_df)] = total_row

                  st.session_state.labor_variance_data = report_df
                  st.session_state.labor_variance_ready = True

                  if export_format == "Excel":
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                              report_df.to_excel(writer, sheet_name="Labor Variance", index=False, startrow=6, header=False)
                              workbook  = writer.book
                              worksheet = writer.sheets["Labor Variance"]

                              # ───── Formats ─────
                              title_fmt       = workbook.add_format({'bold': True, 'font_size': 14})
                              label_fmt       = workbook.add_format({'bold': True})
                              value_fmt       = workbook.add_format({})
                              pos_fmt         = workbook.add_format({'bold': True, 'italic': True, 'border': 1, 'align': 'center'})
                              header_fmt      = workbook.add_format({'bold': True, 'bg_color': '#595959', 'font_color': '#FFFFFF', 'border': 1, 'align': 'center'})
                              default_fmt     = workbook.add_format({'border': 1, 'num_format': '#,##0.00', 'align': 'right'})
                              variance_fmt    = workbook.add_format({'bg_color': '#FEF6F0', 'border': 1, 'num_format': '#,##0.00', 'align': 'right'})
                              total_fmt       = workbook.add_format({'bg_color': '#DCE6F1', 'bold': True, 'border': 1, 'num_format': '#,##0.00', 'align': 'right'})
                              total_text_fmt  = workbook.add_format({'bg_color': '#DCE6F1', 'bold': True, 'italic': True, 'border': 1, 'align': 'center'})
                              total_varpct_fmt= workbook.add_format({'bg_color': '#DCE6F1', 'bold': True, 'border': 1, 'align': 'right'})

                              # ───── Title + Metadata ─────
                              worksheet.write("A1", "Labor Variance Report", title_fmt)
                              worksheet.write("A3", "Department:", label_fmt)
                              worksheet.write("B3", sel_dept, value_fmt)
                              worksheet.write("A4", "Position:", label_fmt)
                              worksheet.write("B4", sel_pos, value_fmt)
                              worksheet.write("A5", "Week:", label_fmt)
                              worksheet.write("B5", f"{week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}", value_fmt)

                              # ───── Header Row ─────
                              for col_idx, col in enumerate(report_df.columns):
                                    worksheet.write(6, col_idx, col, header_fmt)

                              # ───── Precompute TOTAL % Variance ─────
                              total_row_index = report_df[report_df["Position"] == "TOTAL"].index[0]
                              total_actual    = report_df["Actual Hours"].iloc[:-1].sum()
                              total_projected = report_df["Projected Hours"].iloc[:-1].sum()
                              if total_projected:
                                    total_var    = total_actual - total_projected
                                    total_varpct = f"{'▲' if total_var > 0 else '▼' if total_var < 0 else ''} {abs(total_var / total_projected * 100):.2f}%"
                              else:
                                    total_varpct = "–"

                              # ───── Data Rows ─────
                              for row_idx, row in report_df.iterrows():
                                    is_total = row["Position"] == "TOTAL"
                                    for col_idx, col in enumerate(report_df.columns):
                                          val = row[col]

                                          # Inject calculated Variance % for TOTAL row
                                          if is_total and col == "Variance %":
                                                val = total_varpct

                                          # Select format
                                          if col == "Position":
                                                fmt = total_text_fmt if is_total else pos_fmt
                                          elif col == "Variance %":
                                                fmt = total_varpct_fmt if is_total else variance_fmt
                                          elif "Variance" in col:
                                                fmt = total_fmt if is_total else variance_fmt
                                          else:
                                                fmt = total_fmt if is_total else default_fmt

                                          worksheet.write(row_idx + 7, col_idx, val, fmt)

                              worksheet.set_column(0, len(report_df.columns) - 1, 18)

                        st.download_button(
                              "📥 Download Excel",
                              data=output.getvalue(),
                              file_name=f"Labor_Variance_Report_{week_start:%Y%m%d}.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                  elif export_format == "PDF":
                        import io
                        from reportlab.lib.pagesizes import landscape, letter
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib import colors
                        from reportlab.lib.styles import getSampleStyleSheet

                        buffer = io.BytesIO()
                        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
                        styles = getSampleStyleSheet()
                        elements = []

                        # ───── Title and Metadata ─────
                        elements.append(Paragraph("<b>Labor Variance Report</b>", styles["Title"]))
                        elements.append(Spacer(1, 12))
                        elements.append(Paragraph(f"<b>Department:</b> {sel_dept}", styles["Normal"]))
                        elements.append(Paragraph(f"<b>Position:</b> {sel_pos}", styles["Normal"]))
                        elements.append(Paragraph(f"<b>Week:</b> {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}", styles["Normal"]))
                        elements.append(Spacer(1, 12))

                        # ───── Table Data ─────
                        pdf_data = [report_df.columns.tolist()] + report_df.values.tolist()

                        table = Table(pdf_data)

                        style = TableStyle([
                              ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#595959")),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("FONTSIZE", (0, 0), (-1, -1), 9),
                              ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                              ("BACKGROUND", (0, 1), (-1, -2), colors.white),
                              ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ])

                        for i, row in enumerate(report_df.itertuples(index=False), start=1):
                              is_total = row.Position == "TOTAL"
                              if is_total:
                                    style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#DCE6F1"))
                                    style.add("FONTNAME", (0, i), (-1, i), "Helvetica-Bold")
                              else:
                                    style.add("BACKGROUND", (-2, i), (-1, i), colors.HexColor("#FEF6F0"))

                        table.setStyle(style)
                        elements.append(table)
                        doc.build(elements)

                        st.download_button(
                              "📥 Download PDF",
                              data=buffer.getvalue(),
                              file_name=f"Labor_Variance_Report_{week_start:%Y%m%d}.pdf",
                              mime="application/pdf"
                        )

                  elif export_format == "CSV":
                        csv_df = report_df.copy()

                        # Replace arrows with + / - and use 0 when projected hours is 0
                        csv_df["Variance %"] = csv_df.apply(
                              lambda row: f"+ {abs((row['Variance'] / row['Projected Hours']) * 100):.2f}%" if row["Projected Hours"] and row["Variance"] > 0
                              else f"- {abs((row['Variance'] / row['Projected Hours']) * 100):.2f}%" if row["Projected Hours"] and row["Variance"] < 0
                              else "0.00%",
                              axis=1
                        )

                        csv_data = csv_df.to_csv(index=False, encoding="utf-8-sig")

                        st.download_button(
                              "📥 Download CSV",
                              csv_data,
                              file_name=f"Labor_Variance_Report_{week_start:%Y%m%d}.csv",
                              mime="text/csv"
                        )

                  elif export_format == "Excel":
                        import xlsxwriter
                        import io

                        output = io.BytesIO()
                        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
                        worksheet = workbook.add_worksheet("Schedule Variance")

                        # Styles
                        header_fmt = workbook.add_format({
                              'bold': True, 'align': 'center', 'valign': 'vcenter',
                              'bg_color': '#404040', 'font_color': 'white', 'border': 1
                        })
                        text_fmt = workbook.add_format({'align': 'center', 'border': 1})
                        bold_italic_fmt = workbook.add_format({'bold': True, 'italic': True, 'align': 'center', 'border': 1})
                        orange_fmt = workbook.add_format({'bg_color': '#FEF6F0', 'align': 'center', 'border': 1})
                        blue_fmt = workbook.add_format({'bg_color': '#DAEAF3', 'align': 'center', 'border': 1, 'bold': True})
                        blue_bold_fmt = workbook.add_format({'bg_color': '#DAEAF3', 'align': 'center', 'bold': True, 'border': 1})
                        title_fmt = workbook.add_format({'bold': True, 'font_size': 14})
                        meta_key_fmt = workbook.add_format({'bold': True, 'align': 'right'})
                        meta_val_fmt = workbook.add_format({'align': 'left'})

                        # Title & Metadata
                        worksheet.write("A1", "Schedule Variance Report", title_fmt)
                        worksheet.write("A3", "Department:", meta_key_fmt)
                        worksheet.write("B3", sel_dept, meta_val_fmt)
                        worksheet.write("A4", "Week:", meta_key_fmt)
                        worksheet.write("B4", f"{week_start} to {week_end}", meta_val_fmt)

                        # Table headers
                        headers = ["Position", "Scheduled Hours", "Projected Hours", "Variance", "Variance %"]
                        for col, h in enumerate(headers):
                              worksheet.write(6, col, h, header_fmt)

                        # Table rows
                        for row_num, row_data in enumerate(schedule_variance_df.itertuples(index=False), start=7):
                              is_total = row_data.Position == "TOTAL"
                              for col_num, value in enumerate(row_data):
                                    if col_num == 0:
                                          fmt = blue_bold_fmt if is_total else bold_italic_fmt
                                    elif col_num == 3:
                                          fmt = blue_fmt if is_total else orange_fmt
                                    elif col_num == 4:
                                          fmt = blue_fmt if is_total else orange_fmt
                                    else:
                                          fmt = blue_fmt if is_total else text_fmt
                                    worksheet.write(row_num, col_num, value, fmt)

                        worksheet.set_column("A:A", 20)
                        worksheet.set_column("B:E", 18)

                        workbook.close()
                        output.seek(0)
                        st.download_button(
                              label="📥 Download Excel",
                              data=output,
                              file_name=f"Schedule_Variance_{week_start:%Y%m%d}.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

      elif report_type == "Schedule Variance":
            st.markdown("### 📗 Schedule Variance Report")

            from datetime import date, timedelta
            from dateutil.relativedelta import relativedelta, MO
            import pandas as pd
            import io

            if "sched_var_date" not in st.session_state:
                  st.session_state.sched_var_date = date.today()

            sel_date = st.date_input("Select any date in the week", value=st.session_state.sched_var_date, key="sched_var_date")
            week_start = sel_date + relativedelta(weekday=MO(-1))
            week_end   = week_start + timedelta(days=6)

            st.markdown(f"📅 **Selected Week:** {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}")

            emp_df = refresh(db.Employee)
            dept_list = sorted(emp_df["department"].dropna().unique())
            sel_dept = st.selectbox("🏢 Department", dept_list, key="sched_var_dept")

            generate_sched_var = st.button("📊 Generate Schedule Variance Report")
            if generate_sched_var:
                  schedule_df     = refresh(db.Schedule)
                  room_actual_df  = refresh(db.RoomActual)
                  std_df          = refresh(db.LaborStandard)
                  pos_df          = refresh(db.Position).rename(columns={"id": "position_id", "name": "position"})

                  emp_filtered = emp_df[emp_df["department"] == sel_dept]
                  pos_names = emp_filtered["role"].dropna().unique()
                  pos_match = pos_df[pos_df["position"].isin(pos_names)]

                  results = []
                  for _, row in pos_match.iterrows():
                        pos_name = row["position"]
                        pos_id   = row["position_id"]

                        emp_ids = emp_df[
                              (emp_df["department"] == sel_dept) &
                              (emp_df["role"] == pos_name)
                        ]["id"].tolist()

                        sched_rows = schedule_df[
                              (schedule_df["emp_id"].isin(emp_ids)) &
                              (schedule_df["day"] >= week_start) &
                              (schedule_df["day"] <= week_end)
                        ]
                        sched_hours = sched_rows["shift_type"].apply(lambda x: 0 if x == "OFF" else (
                              pd.to_datetime(x.split("-")[1]) - pd.to_datetime(x.split("-")[0])
                        ).seconds / 3600).sum()

                        std_pos = std_df[std_df["position_id"] == pos_id]
                        proj_hours_total = 0
                        for _, std_row in std_pos.iterrows():
                              metric = std_row["metric"]
                              standard = std_row["standard"]
                              if not standard or standard == 0:
                                    continue
                              actual_output = room_actual_df[
                                    (room_actual_df["kpi"] == metric) &
                                    (room_actual_df["date"] >= week_start) &
                                    (room_actual_df["date"] <= week_end)
                              ]["value"].sum()
                              proj_hours_total += (actual_output / standard) * 8

                        projected_hours = proj_hours_total
                        variance = sched_hours - projected_hours
                        variance_pct = (variance / projected_hours * 100) if projected_hours else 0

                        results.append({
                              "Position": pos_name,
                              "Scheduled Hours": round(sched_hours, 1),
                              "Projected Hours": round(projected_hours, 1),
                              "Variance": round(variance, 1),
                              "Variance %": f"{'▲' if variance > 0 else '▼' if variance < 0 else ''} {abs(variance_pct):.2f}%" if projected_hours else "0.00%"
                        })

                  report_df = pd.DataFrame(results)
                  report_df = report_df.groupby("Position", as_index=False).sum(numeric_only=True).sort_values("Position")

                  total_row = {
                        "Position": "TOTAL",
                        "Scheduled Hours": report_df["Scheduled Hours"].sum(),
                        "Projected Hours": report_df["Projected Hours"].sum(),
                        "Variance": report_df["Variance"].sum(),
                        "Variance %": ""
                  }
                  report_df.loc[len(report_df)] = total_row

                  st.session_state.schedule_variance_df = report_df
                  st.session_state.schedule_variance_ready = True

                  if export_format == "Excel" and st.session_state.get("schedule_variance_ready"):
                        report_df = st.session_state.schedule_variance_df.copy()

                        # ───── Recalculate Variance & Variance % ─────
                        report_df["Variance"] = report_df["Scheduled Hours"] - report_df["Projected Hours"]
                        report_df["Variance %"] = report_df.apply(
                              lambda row: f"{'▲' if row['Variance'] > 0 else '▼' if row['Variance'] < 0 else ''} {abs((row['Variance'] / row['Projected Hours']) * 100):.2f}%"
                              if row["Projected Hours"] else "–", axis=1
                        )

                        # ───── Remove any existing TOTAL row ─────
                        report_df = report_df[report_df["Position"] != "TOTAL"]

                        # ───── Add TOTAL Row ─────
                        total_sched = report_df["Scheduled Hours"].sum()
                        total_proj  = report_df["Projected Hours"].sum()
                        total_var   = total_sched - total_proj
                        total_pct   = f"{'▲' if total_var > 0 else '▼' if total_var < 0 else ''} {abs(total_var / total_proj * 100):.2f}%" if total_proj else "–"

                        report_df.loc[len(report_df)] = {
                              "Position": "TOTAL",
                              "Scheduled Hours": round(total_sched, 1),
                              "Projected Hours": round(total_proj, 1),
                              "Variance": round(total_var, 1),
                              "Variance %": total_pct
                        }

                        # ───── Excel Export ─────
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                              report_df.to_excel(writer, sheet_name="Schedule Variance", index=False, startrow=6, header=False)
                              workbook  = writer.book
                              worksheet = writer.sheets["Schedule Variance"]

                              # ───── Formats ─────
                              title_fmt         = workbook.add_format({'bold': True, 'font_size': 14})
                              label_fmt         = workbook.add_format({'bold': True})
                              value_fmt         = workbook.add_format({})
                              pos_fmt           = workbook.add_format({'bold': True, 'italic': True, 'border': 1, 'align': 'center'})
                              header_fmt        = workbook.add_format({'bold': True, 'bg_color': '#595959', 'font_color': '#FFFFFF', 'border': 1, 'align': 'center'})
                              default_fmt       = workbook.add_format({'border': 1, 'num_format': '#,##0.00', 'align': 'right'})
                              variance_fmt      = workbook.add_format({'bg_color': '#FEF6F0', 'border': 1, 'num_format': '#,##0.00', 'align': 'right'})
                              total_fmt         = workbook.add_format({'bg_color': '#DCE6F1', 'bold': True, 'border': 1, 'num_format': '#,##0.00', 'align': 'right'})
                              total_text_fmt    = workbook.add_format({'bg_color': '#DCE6F1', 'bold': True, 'italic': True, 'border': 1, 'align': 'center'})
                              total_varpct_fmt  = workbook.add_format({'bg_color': '#DCE6F1', 'bold': True, 'border': 1, 'align': 'right'})

                              # ───── Title + Metadata ─────
                              worksheet.write("A1", "Schedule Variance Report", title_fmt)
                              worksheet.write("A3", "Department:", label_fmt)
                              worksheet.write("B3", sel_dept, value_fmt)
                              worksheet.write("A4", "Week:", label_fmt)
                              worksheet.write("B4", f"{week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}", value_fmt)

                              # ───── Header Row ─────
                              for col_idx, col in enumerate(report_df.columns):
                                    worksheet.write(6, col_idx, col, header_fmt)

                              # ───── Inject proper % variance for TOTAL ─────
                              total_row_index = report_df[report_df["Position"] == "TOTAL"].index[0]

                              # ───── Data Rows ─────
                              for row_idx, row in report_df.iterrows():
                                    is_total = row["Position"] == "TOTAL"
                                    for col_idx, col in enumerate(report_df.columns):
                                          val = row[col]

                                          # Inject formatted % for total row
                                          if is_total and col == "Variance %":
                                                val = total_pct

                                          # Pick format
                                          if col == "Position":
                                                fmt = total_text_fmt if is_total else pos_fmt
                                          elif col == "Variance %":
                                                fmt = total_varpct_fmt if is_total else variance_fmt
                                          elif "Variance" in col:
                                                fmt = total_fmt if is_total else variance_fmt
                                          else:
                                                fmt = total_fmt if is_total else default_fmt

                                          worksheet.write(row_idx + 7, col_idx, val, fmt)

                              worksheet.set_column(0, len(report_df.columns) - 1, 18)

                        st.download_button(
                              "📥 Download Excel",
                              data=output.getvalue(),
                              file_name=f"Schedule_Variance_Report_{week_start:%Y%m%d}.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                  elif export_format == "PDF" and st.session_state.get("schedule_variance_ready"):
                        report_df = st.session_state.schedule_variance_df.copy()

                        # ───── Calculate Variance Columns ─────
                        report_df["Variance"] = report_df["Scheduled Hours"] - report_df["Projected Hours"]
                        report_df["Variance %"] = report_df.apply(
                              lambda row: f"{'▲' if row['Variance'] > 0 else '▼' if row['Variance'] < 0 else ''} {abs((row['Variance'] / row['Projected Hours']) * 100):.2f}%"
                              if row["Projected Hours"] else "0%", axis=1
                        )

                        # ───── Remove Duplicate TOTAL if exists ─────
                        report_df = report_df[report_df["Position"] != "TOTAL"]

                        # ───── Add Total Row ─────
                        total_sched = report_df["Scheduled Hours"].sum()
                        total_proj  = report_df["Projected Hours"].sum()
                        total_var   = total_sched - total_proj
                        total_pct   = (total_var / total_proj * 100) if total_proj else 0

                        report_df.loc[len(report_df)] = {
                              "Position": "TOTAL",
                              "Scheduled Hours": round(total_sched, 1),
                              "Projected Hours": round(total_proj, 1),
                              "Variance": round(total_var, 1),
                              "Variance %": f"{'▲' if total_var > 0 else '▼' if total_var < 0 else ''} {abs(total_pct):.2f}%"
                        }

                        # ───── PDF Export ─────
                        import io
                        from reportlab.lib.pagesizes import landscape, letter
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors

                        buffer = io.BytesIO()
                        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
                        styles = getSampleStyleSheet()
                        elements = []

                        # Title + Metadata
                        elements.append(Paragraph("Schedule Variance Report", styles["Heading2"]))
                        elements.append(Spacer(1, 6))
                        elements.append(Paragraph(f"<b>Department:</b> {sel_dept}", styles["Normal"]))
                        elements.append(Paragraph(f"<b>Week:</b> {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}", styles["Normal"]))
                        elements.append(Spacer(1, 12))

                        # Table Data
                        table_data = [list(report_df.columns)] + report_df.values.tolist()

                        style = TableStyle([
                              ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#595959")),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                              ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ])

                        for i, row in enumerate(report_df.itertuples(index=False), start=1):
                              if row.Position == "TOTAL":
                                    style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#DCE6F1"))
                                    style.add("FONTNAME", (0, i), (-1, i), "Helvetica-Bold")
                              else:
                                    style.add("BACKGROUND", (3, i), (3, i), colors.HexColor("#FEF6F0"))
                                    style.add("BACKGROUND", (4, i), (4, i), colors.HexColor("#FEF6F0"))

                        table = Table(table_data, repeatRows=1)
                        table.setStyle(style)
                        elements.append(table)

                        doc.build(elements)
                        st.download_button(
                              "📥 Download PDF",
                              data=buffer.getvalue(),
                              file_name=f"Schedule_Variance_Report_{week_start:%Y%m%d}.pdf",
                              mime="application/pdf"
                        )

                  elif export_format == "CSV" and st.session_state.get("schedule_variance_ready"):
                        report_df = st.session_state.schedule_variance_df.copy()

                        # ───── Calculate Variance Columns ─────
                        report_df["Variance"] = report_df["Scheduled Hours"] - report_df["Projected Hours"]
                        report_df["Variance %"] = report_df.apply(
                              lambda row: f"{'+' if row['Variance'] > 0 else '-' if row['Variance'] < 0 else ''} {abs((row['Variance'] / row['Projected Hours']) * 100):.2f}%"
                              if row["Projected Hours"] else "0%", axis=1
                        )

                        # ───── Remove Duplicate TOTAL if exists ─────
                        report_df = report_df[report_df["Position"] != "TOTAL"]

                        # ───── Add Total Row ─────
                        total_sched = report_df["Scheduled Hours"].sum()
                        total_proj  = report_df["Projected Hours"].sum()
                        total_var   = total_sched - total_proj
                        total_pct   = (total_var / total_proj * 100) if total_proj else 0

                        report_df.loc[len(report_df)] = {
                              "Position": "TOTAL",
                              "Scheduled Hours": round(total_sched, 1),
                              "Projected Hours": round(total_proj, 1),
                              "Variance": round(total_var, 1),
                              "Variance %": f"{'+' if total_var > 0 else '-' if total_var < 0 else ''} {abs(total_pct):.2f}%"
                        }

                        # ───── Export as CSV ─────
                        csv_data = report_df.to_csv(index=False)
                        st.download_button(
                              "📥 Download CSV",
                              data=csv_data,
                              file_name=f"Schedule_Variance_Report_{week_start:%Y%m%d}.csv",
                              mime="text/csv"
                        )

if __name__ == "__main__":
    pass
